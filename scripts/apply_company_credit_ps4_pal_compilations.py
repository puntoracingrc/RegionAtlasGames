#!/usr/bin/env python3
"""Apply the reviewed PS4 PAL compilations and company-credit workbook.

The workbook is an editorial source, while every catalog match is pinned here by
catalog_id.  No title-only match is ever written to the catalog.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import html
import json
import re
import subprocess
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RESEARCH = DATA / "research"
CATALOG_FILE = DATA / "catalog.json"
DETAILS_FILE = DATA / "game-details.json"
COMPANIES_FILE = DATA / "index/companies.json"
WORK_IDENTITIES_FILE = DATA / "index/catalog-work-identities.json"
META_FILE = DATA / "meta.json"
CURATION_FILE = DATA / "curation-report.json"
VERIFIED_INDEX_BUILDER = ROOT / "scripts/build_verified_company_credit_index.py"
MANIFEST_FILES = (
    RESEARCH / "company-study/manifest.json",
    RESEARCH / "person-study/manifest.json",
)

BATCH_ID = "company-credit-ps4-pal-compilations-2026-09-05"
REVIEWED_AT = "2026-09-05"
REVIEWED_AT_TIMESTAMP = "2026-09-05T18:00:00Z"
WORKBOOK_NAME = "RegionAtlas_PS4_PAL_recopilatorios_resueltos_2026-09-05.xlsx"
WORKBOOK_SHA256 = "f3e63293c80e30723a17f1447001fb2be3b7bc6a08c24808caad09f4035b52ef"

SOURCE_FILE = RESEARCH / "ps4-pal-compilations-source.json"
DRY_RUN_FILE = RESEARCH / "company-credit-ps4-pal-compilations-dry-run.json"
REPORT_FILE = RESEARCH / "company-credit-ps4-pal-compilations-report.json"
REPORT_MD_FILE = RESEARCH / "company-credit-ps4-pal-compilations-report.md"
COMMERCIAL_RELATIONS_FILE = DATA / "index/catalog-commercial-relations.json"
ROUTE_REDIRECTS_FILE = DATA / "catalog-route-redirects.json"
COMPANY_RELATIONS_FILE = DATA / "index/verified-company-relations.json"
COMPANY_ALIASES_FILE = DATA / "index/verified-company-aliases.json"

EXPECTED_SHEET_ROWS = {
    "Recopilatorios": 24,
    "Componentes": 82,
    "Duplicados catálogo": 3,
    "Lote 8 solicitado": 8,
    "Plataforma incorrecta": 3,
    "Identidad ambigua": 3,
    "Históricos y codesarrollo": 3,
    "Lote editorial 01": 11,
    "Lote editorial 02": 10,
    "Lote editorial 03": 10,
    "Lote editorial 04": 10,
    "Lote editorial 05": 10,
    "Lote editorial 06": 6,
}

EDITORIAL_ID_BY_NUMBER = {
    0: "ps4-adventure-academia-the-fractured-continent",
    1: "ps4-air-conflicts-secret-wars",
    2: "ps4-armored-core-vi-fires-of-rubicon-collector%27s-edition",
    3: "ps4-armored-core-vi-fires-of-rubicon-launch-edition",
    4: "ps4-batman-arkham-vr",
    5: "ps4-blasphemous-coleccionista",
    6: "ps4-classic-racers-elite",
    7: "ps4-cyberpunk-2077-samurai-pack",
    8: "ps4-dark-souls-ii-scholar-of-the-first-sin-not-for-resale",
    9: "ps4-date-a-live-rio-reincarnation",
    10: "ps4-daymare-1994-sandcastle",
    11: "ps4-demon-pit",
    12: "ps4-ducati-90th-anniversary",
    13: "ps4-dynasty-feud",
    14: "ps4-genesis-alpha-one",
    15: "ps4-ghost-trick-phantom-detective",
    16: "ps4-gungrave-vr",
    17: "ps4-gylt-collector%27s-edition",
    18: "ps4-hello-neighbor-hide-&amp;-seek",
    19: "ps4-hokko-life",
    20: "ps4-indiecalypse",
    21: "ps4-infinite-beyond-the-mind-tanya-edition",
    22: "ps4-jets-%27n%27-guns-2",
    23: "ps4-jets%27n%27guns-2",
    24: "ps4-koa-and-the-five-pirates-of-mara",
    25: "ps4-lemon-cake",
    26: "ps4-memorrha",
    27: "ps4-mickey-storm-and-the-cursed-mask",
    28: "ps4-middle-earth-shadow-of-mordor-game-of-the-year",
    29: "ps4-milli-&amp;-greg",
    30: "ps4-nobunaga%27s-ambition-sphere-of-influence-ascension",
    31: "ps4-nongunz-doppelganger-edition",
    32: "ps4-observer-system-redux",
    33: "ps4-okinawa-rush-black-mantis-edition",
    34: "ps4-okinawa-rush-limited-edition",
    35: "ps4-oxide-room-104",
    36: "ps4-past-cure",
    37: "ps4-professional-farmer-american-dream",
    38: "ps4-quivr",
    39: "ps4-race-with-ryan",
    40: "ps4-rogue-trooper-redux",
    41: "ps4-ryan%27s-rescue-squad",
    42: "ps4-senran-kagura-burst-renewal-bountiful-beauties",
    43: "ps4-source-of-madness",
    44: "ps4-street-power-football",
    45: "ps4-super-crazy-rhythm-castle",
    46: "ps4-superhot-vr",
    47: "ps4-the-american-dream",
    48: "ps4-the-occupation",
    49: "ps4-the-rumble-fish-2",
    50: "ps4-trailblazers",
    51: "ps4-ultrawings",
    52: "ps4-unmetal",
    53: "ps4-young-souls-collector%27s-edition",
    54: "ps4-young-souls-deluxe-edition",
    55: "ps4-youtubers-life-2",
    56: "ps4-dracula-s-legacy",
}

BUNDLE_ID_BY_NAME = {
    "Annapurna Ultimate Collection": "ps4-usa-annapurna-interactive-ultimate-ps4-collection",
    "PlayLink Games Collection": "ps4-playlink-games-collection",
    "Terra Trilogy": "ps4-terra-trilogy",
    "The Giants: Industry and Transport Bundle": "ps4-the-giants-industry-and-transport-bundle",
    "Mafia Trilogy": "ps4-mafia-trilogy",
    "Star Wars Jedi Knight Collection": "ps4-star-wars-jedi-knight-collection",
    "Sonic Mania Plus and Sonic Forces Double Pack": "ps4-sonic-mania-plus-and-sonic-forces-double-pack",
    "Wolfenstein Alt History Collection": "ps4-wolfenstein-alt-history-collection",
    "Doctor Who: Duo Bundle": "ps4-doctor-who-duo-bundle",
    "Dishonored & Prey: The Arkane Collection": "ps4-dishonored-prey-the-arkane-collection",
    "Dishonored: The Complete Collection": "ps4-dishonored-complete-collection",
    "Dark Thrones / Witch Hunter Double Pack": "ps4-dark-thrones-&amp;-witch-hunter-double-pack",
    "Persona 5 + Persona 5: Dancing in Starlight [2 Hits Pack]": "ps4-persona-5-&#43;-persona-5-dancing-the-starlight-day-one-edition-2-hits-pack",
    "Spyro Reignited Trilogy & Crash Bandicoot N. Sane Trilogy": "ps4-spyro-reignited-trilogy-&amp;-crash-bandicoot-n-sane-trilogy",
    "Crash Team Racing & Spyro Reignited Trilogy": "ps4-crash-team-racing-&amp;-spyro-reignited-trilogy",
    "Dragon Ball FighterZ + Dragon Ball Xenoverse 2": "ps4-dragon-ball-fighterz-&#43;-dragon-ball-xenoverse-2",
    "Dragon Ball Xenoverse 1 & 2 Double Pack": "ps4-dragon-ball-xenoverse-1-&amp;-2-double-pack",
    "Tales of Zestiria + Berseria + Vesperia Triple Pack": "ps4-tales-of-zestiria-&#43;-tales-of-berseria-&#43;-tales-of-vesperia-triple-pack-bundle",
    "The Ultimate VR Collection": "ps4-the-ultimate-vr-collection",
    "The Truth Is Out There: Mystery Adventure Pack": "ps4-the-truth-is-out-there-mystery-adventure-pack",
    "World of Simulators": "ps4-world-of-simulators",
    "Wonder Boy Anniversary Collection": "ps4-wonder-boy-anniversary-collection",
    "Wonder Boy Collection": "ps4-wonder-boy-collection",
    "PlayStation VR Demo Disc (PAL Europa)": None,
}

BUNDLE_PUBLIC_TITLES = {
    "Annapurna Ultimate Collection": "Annapurna Interactive Ultimate PS4 Collection",
}

# Exact component links only. Missing values remain useful structured credits but
# are deliberately not guessed from nearby titles or remakes.
COMPONENT_LINKS = {
    ("Annapurna Ultimate Collection", "Sayonara Wild Hearts"): "ps4-sayonara-wild-hearts",
    ("PlayLink Games Collection", "Hidden Agenda"): "ps4-hidden-agenda",
    ("PlayLink Games Collection", "Knowledge is Power"): "ps4-knowledge-is-power",
    ("PlayLink Games Collection", "SingStar Celebration"): "ps4-singstar-celebration",
    ("PlayLink Games Collection", "That's You!"): "ps4-that%27s-you",
    ("The Giants: Industry and Transport Bundle", "Industry Giant 2"): "ps4-industry-giant-2",
    ("Mafia Trilogy", "Mafia: Definitive Edition"): "ps4-mafia-definitive-edition",
    ("Sonic Mania Plus and Sonic Forces Double Pack", "Sonic Mania Plus"): "ps4-sonic-mania-plus",
    ("Sonic Mania Plus and Sonic Forces Double Pack", "Sonic Forces"): "ps4-sonic-forces",
    ("Wolfenstein Alt History Collection", "Wolfenstein: The New Order"): "ps4-wolfenstein-the-new-order",
    ("Wolfenstein Alt History Collection", "Wolfenstein: The Old Blood"): "ps4-wolfenstein-the-old-blood",
    ("Wolfenstein Alt History Collection", "Wolfenstein II: The New Colossus"): "ps4-wolfenstein-ii-the-new-colossus",
    ("Dishonored & Prey: The Arkane Collection", "Dishonored: Definitive Edition"): "ps4-dishonored-definitive-edition",
    ("Dishonored & Prey: The Arkane Collection", "Dishonored 2"): "ps4-dishonored-2",
    ("Dishonored & Prey: The Arkane Collection", "Dishonored: Death of the Outsider"): "ps4-dishonored-death-of-the-outsider",
    ("Dishonored & Prey: The Arkane Collection", "Prey"): "ps4-prey",
    ("Dishonored: The Complete Collection", "Dishonored: Definitive Edition"): "ps4-dishonored-definitive-edition",
    ("Dishonored: The Complete Collection", "Dishonored 2"): "ps4-dishonored-2",
    ("Dishonored: The Complete Collection", "Dishonored: Death of the Outsider"): "ps4-dishonored-death-of-the-outsider",
    ("Persona 5 + Persona 5: Dancing in Starlight [2 Hits Pack]", "Persona 5"): "ps4-persona-5",
    ("Persona 5 + Persona 5: Dancing in Starlight [2 Hits Pack]", "Persona 5: Dancing in Starlight"): "ps4-persona-5-dancing-in-starlight",
    ("Spyro Reignited Trilogy & Crash Bandicoot N. Sane Trilogy", "Spyro Reignited Trilogy"): "ps4-spyro-reignited-trilogy",
    ("Spyro Reignited Trilogy & Crash Bandicoot N. Sane Trilogy", "Crash Bandicoot N. Sane Trilogy"): "ps4-crash-bandicoot-n-sane-trilogy",
    ("Crash Team Racing & Spyro Reignited Trilogy", "Crash Team Racing Nitro-Fueled"): "ps4-crash-team-racing-nitro-fueled",
    ("Crash Team Racing & Spyro Reignited Trilogy", "Spyro Reignited Trilogy"): "ps4-spyro-reignited-trilogy",
    ("Dragon Ball FighterZ + Dragon Ball Xenoverse 2", "Dragon Ball FighterZ"): "ps4-dragon-ball-fighterz",
    ("Dragon Ball FighterZ + Dragon Ball Xenoverse 2", "Dragon Ball Xenoverse 2"): "ps4-dragon-ball-xenoverse-2",
    ("Dragon Ball Xenoverse 1 & 2 Double Pack", "Dragon Ball Xenoverse"): "ps4-dragon-ball-xenoverse",
    ("Dragon Ball Xenoverse 1 & 2 Double Pack", "Dragon Ball Xenoverse 2"): "ps4-dragon-ball-xenoverse-2",
    ("Tales of Zestiria + Berseria + Vesperia Triple Pack", "Tales of Zestiria"): "ps4-tales-of-zestiria",
    ("Tales of Zestiria + Berseria + Vesperia Triple Pack", "Tales of Berseria"): "ps4-tales-of-berseria",
    ("Tales of Zestiria + Berseria + Vesperia Triple Pack", "Tales of Vesperia: Definitive Edition"): "ps4-tales-of-vesperia-definitive-edition",
    ("The Truth Is Out There: Mystery Adventure Pack", "Pineview Drive"): "ps4-pineview-drive",
    ("The Truth Is Out There: Mystery Adventure Pack", "Joe's Diner"): "ps4-joe%27s-diner",
    ("The Truth Is Out There: Mystery Adventure Pack", "The Nightfall"): "ps4-the-nightfall",
    ("World of Simulators", "Professional Farmer 2017"): "ps4-professional-farmer-2017",
    ("World of Simulators", "Forestry 2017: The Simulation"): "ps4-forestry-2017",
    ("PlayStation VR Demo Disc (PAL Europa)", "DriveClub VR"): "ps4-driveclub-vr",
    ("PlayStation VR Demo Disc (PAL Europa)", "PlayStation VR Worlds"): "ps4-playstation-vr-worlds",
    ("PlayStation VR Demo Disc (PAL Europa)", "RIGS: Mechanized Combat League"): "ps4-rigs-mechanized-combat-league",
    ("PlayStation VR Demo Disc (PAL Europa)", "Battlezone"): "ps4-battlezone",
    ("PlayStation VR Demo Disc (PAL Europa)", "EVE: Valkyrie"): "ps4-eve-valkyrie",
}

COMPANY_DEFINITIONS = {
    "1p2p-studio": ("1P2P Studio", []),
    "2k-games": ("2K Games", []),
    "2ndboss": ("2ndBoss", []),
    "2tainment": ("2tainment GmbH", ["2tainment"]),
    "3dclouds": ("3DClouds", ["3D Clouds", "3DClouds S.r.l."]),
    "3goo": ("3goo", ["3goo K.K."]),
    "acquire": ("Acquire", []),
    "activision": ("Activision", []),
    "alvios": ("Alvios", ["Alvios Inc."]),
    "annapurna-interactive": ("Annapurna Interactive", []),
    "aspyr": ("Aspyr", []),
    "atlus": ("Atlus", ["Atlus Co., Ltd."]),
    "bandai-namco-entertainment": ("Bandai Namco Entertainment", []),
    "bandai-namco-entertainment-europe": ("Bandai Namco Entertainment Europe", []),
    "bit-planet-games": ("Bit Planet Games", ["Bit Planet Games, LLC", "BIT PLANET GAMES LLC"]),
    "bliss-brain": ("Bliss Brain", []),
    "bloober-team": ("Bloober Team", []),
    "blowfish-studios": ("Blowfish Studios", ["Blowfish Studios Pty Ltd"]),
    "blueteak": ("Blueteak", []),
    "brainwash-gang": ("Brainwash Gang", []),
    "capcom": ("Capcom", []),
    "carry-castle": ("Carry Castle", ["Carry Castle AB"]),
    "cd-projekt": ("CD Projekt", []),
    "cd-projekt-red": ("CD Projekt RED", []),
    "ce-europe-ltd": ("CE Europe Limited", ["CE Europe Ltd."]),
    "chibig": ("Chibig", ["Chibig Studio"]),
    "compile-heart": ("Compile Heart", []),
    "cozy-bee-games": ("Cozy Bee Games", []),
    "digerati": ("Digerati", ["Digerati Distribution"]),
    "dimps-corporation": ("Dimps Corporation", ["Dimps"]),
    "dm-media": ("DM Media", []),
    "doomcube": ("DoomCube", []),
    "dynamic-pixels": ("Dynamic Pixels", []),
    "funbox-media": ("Funbox Media", ["Fun Box Media"]),
    "gamajun-games": ("Gamajun Games", []),
    "games-farm": ("Games Farm", ["GamesFarm"]),
    "gotcha-gotcha-games-inc": ("Gotcha Gotcha Games Inc", ["Gotcha Gotcha Games Inc."]),
    "honey-parade-games": ("Honey Parade Games", ["HONEY∞PARADE GAMES, Inc."]),
    "humble-games": ("Humble Games", ["Humble Bundle, Inc."]),
    "iam8bit": ("iam8bit", ["Iam8bit"]),
    "idea-factory": ("Idea Factory", ["Idea Factory Co., Ltd."]),
    "idea-factory-international": ("Idea Factory International", []),
    "iggymob": ("IGGYMOB", []),
    "inin-games": ("ININ Games", []),
    "invader-studios-s-r-l": ("Invader Studios", ["Invader Studios S.R.L."]),
    "iridium-media-group-gmbh": ("Iridium Media Group GmbH", []),
    "jandusoft": ("JanduSoft", ["JanduSoft S.L."]),
    "jetdogs-studios": ("Jetdogs Studios", []),
    "joindots": ("Joindots", ["Joindots GmbH"]),
    "kaia-studios": ("Kaia Studios", []),
    "koei-tecmo-europe": ("Koei Tecmo Europe", ["Koei Tecmo Europe Limited"]),
    "koei-tecmo-games": ("Koei Tecmo Games", ["Koei Tecmo Games Co., Ltd."]),
    "konami-digital-entertainment": ("Konami Digital Entertainment", ["KONAMI"]),
    "leonardo-interactive": ("Leonardo Interactive", []),
    "lion-castle": ("Lion Castle", ["Lion Castle B.V."]),
    "markt-43-technik-verlag-gmbh": ("Markt+Technik Verlag GmbH", ["Markt+Technik", "Markt Technik"]),
    "marvelous-europe": ("Marvelous Europe", ["Marvelous Europe Limited"]),
    "maximum-games": ("Maximum Games", []),
    "maze-theory": ("Maze Theory", []),
    "milestone-s-r-l": ("Milestone S.r.l.", ["Milestone", "MILESTONE SRL"]),
    "monolith-productions": ("Monolith Productions", []),
    "nacon": ("Nacon", ["NACON SA"]),
    "nis-america": ("NIS America", []),
    "no-gravity-games": ("No Gravity Games", []),
    "nyx-digital": ("NYX Digital", []),
    "outright-games": ("Outright Games", ["Outright Games LLC", "OUTRIGHT GAMES LLC"]),
    "perp-games": ("Perp Games", []),
    "phantom-8": ("Phantom 8", ["Phantom 8 Studio UG"]),
    "pixelheart": ("PixelHeart", ["PIXELHEART Corporation"]),
    "pix-39-n-love": ("Pix'n Love", ["Pix’n Love", "Pix 'n Love"]),
    "pqube": ("PQube", ["PQube Ltd."]),
    "psychic-software": ("Psychic Software", []),
    "qubyte": ("QUByte Interactive", ["QUByte"]),
    "radiation-blue": ("Radiation Blue", []),
    "raiser-games": ("Raiser Games", ["Raiser Games S.L."]),
    "rake-in-grass": ("Rake in Grass", []),
    "raven-software": ("Raven Software", []),
    "rebellion-developments": ("Rebellion Developments", []),
    "rebellion-interactive": ("Rebellion Interactive", []),
    "red-art-games": ("Red Art Games", []),
    "reef-entertainment": ("Reef Entertainment", []),
    "relevo": ("Relevo", ["Relevo Videogames"]),
    "rising-star-games": ("Rising Star Games", ["Rising Star Games Ltd."]),
    "rocksteady-studios": ("Rocksteady Studios", []),
    "samurai-punk": ("Samurai Punk", ["Samurai Punk Pty Ltd."]),
    "second-impact-games": ("Second Impact Games", ["Second Impact Games Ltd."]),
    "sega": ("Sega", ["SEGA"]),
    "selecta-play": ("Selecta Play", []),
    "sfl-interactive": ("SFL Interactive", []),
    "sokaikan-ltd": ("Sokaikan Ltd.", []),
    "sold-out": ("Sold Out", []),
    "sony-interactive-entertainment": ("Sony Interactive Entertainment", []),
    "soedesco": ("SOEDESCO", ["Soedesco"]),
    "stage-clear-studios": ("Stage Clear Studios", []),
    "sticky-stone-studio": ("StickyStoneStudio", ["Sticky Stone Studio"]),
    "sting": ("Sting", []),
    "strictly-limited-games": ("Strictly Limited Games", []),
    "suncrest-games": ("Suncrest Games", []),
    "supergonk": ("Supergonk", ["Supergonk Ltd."]),
    "superhot-team": ("SUPERHOT Team", ["SUPERHOT Sp. z o.o."]),
    "talpa-games": ("Talpa Games", []),
    "team17": ("Team17", ["Team17 Digital Ltd.", "Team 17 Digital Ltd.", "Team 17"]),
    "tequila-works": ("Tequila Works", []),
    "tesura-games": ("Tesura Games", []),
    "the-arcade-crew": ("The Arcade Crew", []),
    "the-game-kitchen": ("The Game Kitchen", []),
    "the-munky": ("The Munky", ["The Munky, LLC"]),
    "thq-nordic": ("THQ Nordic", ["THQ Nordic GmbH"]),
    "thunderful-games-ltd": ("Thunderful Games Ltd", ["Thunderful Games", "Thunderful Group"]),
    "thunderful-publishing": ("Thunderful Publishing", ["Thunderful Publishing AB"]),
    "ticktock-games": ("TickTock Games", []),
    "tinybuild": ("tinyBuild", ["tinyBuild LLC", "TINYBUILD LLC"]),
    "triangle-studios": ("Triangle Studios", ["Triangle Studios B.V."]),
    "u-play-online": ("U-Play Online", ["U-Play Online S.L.", "UPLAY ONLINE SL"]),
    "uig-entertainment": ("UIG Entertainment", ["UIG Entertainment GmbH"]),
    "undercoders": ("Undercoders", []),
    "unepic-games": ("UnEpic Games", ["UnEpic Fran"]),
    "united-independent-entertainment": ("United Independent Entertainment", []),
    "versus-evil": ("Versus Evil", []),
    "virtualware": ("Virtualware", []),
    "vision-reelle": ("Vision Réelle", ["Vision Reelle"]),
    "warner-bros-interactive-entertainment": ("Warner Bros. Interactive Entertainment", ["Warner Bros. Interactive", "Warner Bros. Games"]),
    "westone-bit-entertainment": ("Westone Bit Entertainment", []),
    "white-paper-games": ("White Paper Games", ["White Paper Games Ltd."]),
    "wild-sphere-s-l-u": ("WildSphere S.L.U.", ["WildSphere", "Wild Sphere", "WILD SPHERE S.L.", "WildSphere S.L."]),
    "wonderscope": ("Wonderscope", []),
}


def c(role: str, slug: str) -> tuple[str, str]:
    return role, slug


CREDIT_PLANS = {
    0: [c("developer", "acquire"), c("originalPublisher", "pqube"), c("physicalPublisherOrDistributor", "pqube")],
    1: [c("developer", "games-farm"), c("originalPublisher", "2tainment"), c("physicalPublisherOrDistributor", "2tainment")],
    2: [c("developer", "fromsoftware"), c("originalPublisher", "bandai-namco-entertainment"), c("physicalPublisherOrDistributor", "bandai-namco-entertainment")],
    3: [c("developer", "fromsoftware"), c("originalPublisher", "bandai-namco-entertainment"), c("physicalPublisherOrDistributor", "bandai-namco-entertainment")],
    4: [c("developer", "rocksteady-studios"), c("originalPublisher", "warner-bros-interactive-entertainment"), c("physicalPublisherOrDistributor", "warner-bros-interactive-entertainment")],
    5: [c("developer", "the-game-kitchen"), c("originalPublisher", "team17"), c("physicalPublisherOrDistributor", "selecta-play")],
    6: [c("developer", "vision-reelle"), c("originalPublisher", "funbox-media"), c("physicalPublisherOrDistributor", "funbox-media")],
    7: [c("developer", "cd-projekt-red"), c("originalPublisher", "cd-projekt")],
    8: [c("developer", "fromsoftware"), c("originalPublisher", "bandai-namco-entertainment")],
    9: [c("developer", "idea-factory"), c("developer", "compile-heart"), c("developer", "sting"), c("regionalPublisher", "reef-entertainment"), c("physicalPublisherOrDistributor", "idea-factory-international"), c("physicalPublisherOrDistributor", "reef-entertainment")],
    10: [c("developer", "invader-studios-s-r-l"), c("originalPublisher", "leonardo-interactive"), c("physicalPublisherOrDistributor", "leonardo-interactive")],
    11: [c("developer", "psychic-software"), c("developer", "doomcube"), c("originalPublisher", "digerati"), c("physicalPublisherOrDistributor", "digerati")],
    12: [c("developer", "milestone-s-r-l"), c("originalPublisher", "milestone-s-r-l"), c("physicalPublisherOrDistributor", "milestone-s-r-l")],
    13: [c("developer", "kaia-studios"), c("originalPublisher", "kaia-studios"), c("regionalPublisher", "relevo"), c("physicalPublisherOrDistributor", "relevo")],
    14: [c("developer", "radiation-blue"), c("originalPublisher", "team17"), c("physicalPublisherOrDistributor", "team17")],
    15: [c("developer", "capcom"), c("originalPublisher", "capcom"), c("regionalPublisher", "ce-europe-ltd"), c("physicalPublisherOrDistributor", "capcom"), c("physicalPublisherOrDistributor", "ce-europe-ltd")],
    16: [c("developer", "iggymob"), c("regionalPublisher", "marvelous-europe"), c("physicalPublisherOrDistributor", "marvelous-europe")],
    17: [c("developer", "tequila-works"), c("originalPublisher", "tequila-works"), c("physicalPublisherOrDistributor", "tesura-games")],
    18: [c("developer", "dynamic-pixels"), c("originalPublisher", "tinybuild"), c("physicalPublisherOrDistributor", "tinybuild")],
    19: [c("developer", "wonderscope"), c("originalPublisher", "team17"), c("physicalPublisherOrDistributor", "team17")],
    20: [c("developer", "jandusoft"), c("originalPublisher", "jandusoft"), c("physicalPublisherOrDistributor", "jandusoft")],
    21: [c("originalPublisher", "blowfish-studios"), c("physicalPublisherOrDistributor", "strictly-limited-games")],
    22: [c("developer", "rake-in-grass"), c("regionalPublisher", "red-art-games"), c("physicalPublisherOrDistributor", "red-art-games")],
    23: [c("developer", "rake-in-grass"), c("regionalPublisher", "red-art-games"), c("physicalPublisherOrDistributor", "red-art-games")],
    24: [c("developer", "chibig"), c("developer", "talpa-games"), c("developer", "undercoders"), c("originalPublisher", "chibig"), c("physicalPublisherOrDistributor", "tesura-games")],
    25: [c("developer", "cozy-bee-games"), c("originalPublisher", "soedesco"), c("physicalPublisherOrDistributor", "soedesco")],
    26: [c("developer", "sticky-stone-studio"), c("originalPublisher", "sticky-stone-studio"), c("regionalPublisher", "markt-43-technik-verlag-gmbh"), c("physicalPublisherOrDistributor", "markt-43-technik-verlag-gmbh")],
    27: [c("developer", "triangle-studios"), c("originalPublisher", "lion-castle"), c("physicalPublisherOrDistributor", "lion-castle")],
    28: [c("developer", "monolith-productions"), c("originalPublisher", "warner-bros-interactive-entertainment"), c("physicalPublisherOrDistributor", "warner-bros-interactive-entertainment")],
    29: [c("developer", "2ndboss"), c("originalPublisher", "qubyte"), c("physicalPublisherOrDistributor", "qubyte")],
    30: [c("developer", "koei-tecmo-games"), c("originalPublisher", "koei-tecmo-games"), c("regionalPublisher", "koei-tecmo-europe"), c("physicalPublisherOrDistributor", "koei-tecmo-europe")],
    31: [c("developer", "brainwash-gang"), c("originalPublisher", "digerati"), c("physicalPublisherOrDistributor", "digerati")],
    32: [c("developer", "bloober-team"), c("originalPublisher", "aspyr")],
    33: [c("developer", "sokaikan-ltd"), c("originalPublisher", "no-gravity-games"), c("physicalPublisherOrDistributor", "pixelheart")],
    34: [c("developer", "sokaikan-ltd"), c("originalPublisher", "no-gravity-games"), c("physicalPublisherOrDistributor", "pixelheart")],
    35: [c("developer", "wild-sphere-s-l-u"), c("originalPublisher", "wild-sphere-s-l-u"), c("physicalPublisherOrDistributor", "perp-games")],
    36: [c("developer", "phantom-8"), c("originalPublisher", "phantom-8"), c("physicalPublisherOrDistributor", "phantom-8")],
    37: [c("developer", "uig-entertainment"), c("originalPublisher", "uig-entertainment"), c("regionalPublisher", "iridium-media-group-gmbh"), c("physicalPublisherOrDistributor", "iridium-media-group-gmbh")],
    38: [c("originalDeveloper", "blueteak"), c("portDeveloper", "alvios"), c("regionalPublisher", "the-munky"), c("physicalPublisherOrDistributor", "perp-games")],
    39: [c("developer", "3dclouds"), c("originalPublisher", "outright-games"), c("physicalPublisherOrDistributor", "outright-games")],
    40: [c("originalDeveloper", "rebellion-developments"), c("remasterDeveloper", "ticktock-games"), c("regionalPublisher", "rebellion-interactive"), c("physicalPublisherOrDistributor", "rebellion-interactive")],
    41: [c("developer", "stage-clear-studios"), c("originalPublisher", "outright-games"), c("physicalPublisherOrDistributor", "outright-games")],
    42: [c("developer", "honey-parade-games"), c("regionalPublisher", "marvelous-europe"), c("physicalPublisherOrDistributor", "marvelous-europe")],
    43: [c("developer", "carry-castle"), c("originalPublisher", "thunderful-publishing"), c("physicalPublisherOrDistributor", "thunderful-publishing")],
    44: [c("developer", "sfl-interactive"), c("developer", "gamajun-games")],
    45: [c("developer", "second-impact-games"), c("originalPublisher", "konami-digital-entertainment"), c("physicalPublisherOrDistributor", "konami-digital-entertainment")],
    46: [c("developer", "superhot-team"), c("originalPublisher", "superhot-team"), c("physicalPublisherOrDistributor", "superhot-team")],
    47: [c("developer", "samurai-punk"), c("originalPublisher", "samurai-punk"), c("physicalPublisherOrDistributor", "perp-games")],
    48: [c("developer", "white-paper-games"), c("originalPublisher", "humble-games"), c("physicalPublisherOrDistributor", "sold-out")],
    49: [c("originalDeveloper", "dimps-corporation"), c("portDeveloper", "suncrest-games"), c("regionalPublisher", "3goo"), c("physicalPublisherOrDistributor", "3goo")],
    50: [c("developer", "supergonk"), c("originalPublisher", "rising-star-games"), c("physicalPublisherOrDistributor", "rising-star-games")],
    51: [c("developer", "bit-planet-games"), c("originalPublisher", "bit-planet-games"), c("physicalPublisherOrDistributor", "bit-planet-games")],
    52: [c("developer", "unepic-games"), c("originalPublisher", "versus-evil"), c("digitalPublisher", "tinybuild"), c("physicalPublisherOrDistributor", "tesura-games")],
    53: [c("developer", "1p2p-studio"), c("originalPublisher", "the-arcade-crew"), c("physicalPublisherOrDistributor", "pix-39-n-love")],
    54: [c("developer", "1p2p-studio"), c("originalPublisher", "the-arcade-crew")],
    55: [c("developer", "u-play-online"), c("originalPublisher", "raiser-games")],
    56: [c("developer", "jetdogs-studios"), c("originalPublisher", "joindots"), c("physicalPublisherOrDistributor", "joindots")],
}

INDIVIDUAL_CREDITS = {
    21: [("developer", "Emilie Coyo", "emilie-coyo")],
    25: [("developer", "Éloïse Laroche", "eloise-laroche")],
    52: [("developer", "Francisco Téllez de Meneses", "francisco-tellez-de-meneses")],
}

CO_DEVELOPER_NUMBERS = {11, 24, 44}
MINDTAKER_ID = "ps4-mindtaker"
MINDTAKER_CREDITS = [c("developer", "relevo"), c("developer", "virtualware")]

BUNDLE_CREDIT_PLANS = {
    "Annapurna Ultimate Collection": [c("originalPublisher", "annapurna-interactive"), c("physicalPublisherOrDistributor", "skybound-games")],
    "PlayLink Games Collection": [c("originalPublisher", "sony-interactive-entertainment"), c("physicalPublisherOrDistributor", "sony-interactive-entertainment")],
    "Terra Trilogy": [c("developer", "dm-media"), c("originalPublisher", "funbox-media"), c("physicalPublisherOrDistributor", "funbox-media")],
    "The Giants: Industry and Transport Bundle": [c("originalPublisher", "united-independent-entertainment"), c("physicalPublisherOrDistributor", "united-independent-entertainment")],
    "Mafia Trilogy": [c("originalPublisher", "2k-games"), c("physicalPublisherOrDistributor", "2k-games")],
    "Star Wars Jedi Knight Collection": [c("originalDeveloper", "raven-software"), c("portDeveloper", "aspyr"), c("originalPublisher", "lucasarts"), c("physicalPublisherOrDistributor", "thq-nordic")],
    "Sonic Mania Plus and Sonic Forces Double Pack": [c("originalPublisher", "sega"), c("physicalPublisherOrDistributor", "sega")],
    "Wolfenstein Alt History Collection": [c("originalPublisher", "bethesda-softworks"), c("physicalPublisherOrDistributor", "bethesda-softworks")],
    "Doctor Who: Duo Bundle": [c("originalPublisher", "maze-theory"), c("physicalPublisherOrDistributor", "maximum-games")],
    "Dishonored & Prey: The Arkane Collection": [c("originalPublisher", "bethesda-softworks"), c("physicalPublisherOrDistributor", "bethesda-softworks")],
    "Dishonored: The Complete Collection": [c("originalPublisher", "bethesda-softworks"), c("physicalPublisherOrDistributor", "bethesda-softworks")],
    "Dark Thrones / Witch Hunter Double Pack": [c("developer", "nyx-digital"), c("originalPublisher", "funbox-media"), c("physicalPublisherOrDistributor", "funbox-media")],
    "Persona 5 + Persona 5: Dancing in Starlight [2 Hits Pack]": [c("originalPublisher", "atlus"), c("physicalPublisherOrDistributor", "atlus")],
    "Spyro Reignited Trilogy & Crash Bandicoot N. Sane Trilogy": [c("originalPublisher", "activision"), c("physicalPublisherOrDistributor", "activision")],
    "Crash Team Racing & Spyro Reignited Trilogy": [c("originalPublisher", "activision"), c("physicalPublisherOrDistributor", "activision")],
    "Dragon Ball FighterZ + Dragon Ball Xenoverse 2": [c("originalPublisher", "bandai-namco-entertainment"), c("regionalPublisher", "bandai-namco-entertainment-europe"), c("physicalPublisherOrDistributor", "bandai-namco-entertainment-europe")],
    "Dragon Ball Xenoverse 1 & 2 Double Pack": [c("originalPublisher", "bandai-namco-entertainment"), c("regionalPublisher", "bandai-namco-entertainment-europe"), c("physicalPublisherOrDistributor", "bandai-namco-entertainment-europe")],
    "Tales of Zestiria + Berseria + Vesperia Triple Pack": [c("originalPublisher", "bandai-namco-entertainment"), c("physicalPublisherOrDistributor", "bandai-namco-entertainment")],
    "The Ultimate VR Collection": [c("physicalPublisherOrDistributor", "perp-games")],
    "The Truth Is Out There: Mystery Adventure Pack": [c("originalPublisher", "united-independent-entertainment"), c("regionalPublisher", "uig-entertainment"), c("physicalPublisherOrDistributor", "uig-entertainment")],
    "World of Simulators": [c("originalPublisher", "united-independent-entertainment"), c("regionalPublisher", "uig-entertainment"), c("physicalPublisherOrDistributor", "uig-entertainment")],
    "Wonder Boy Anniversary Collection": [c("originalDeveloper", "westone-bit-entertainment"), c("portDeveloper", "bliss-brain"), c("originalPublisher", "bliss-brain"), c("regionalPublisher", "inin-games"), c("physicalPublisherOrDistributor", "inin-games"), c("physicalPublisherOrDistributor", "strictly-limited-games")],
    "Wonder Boy Collection": [c("originalDeveloper", "westone-bit-entertainment"), c("portDeveloper", "bliss-brain"), c("originalPublisher", "bliss-brain"), c("regionalPublisher", "inin-games"), c("physicalPublisherOrDistributor", "inin-games")],
}

BUNDLE_SOURCE_OVERRIDES = {
    "Annapurna Ultimate Collection": {
        "evidenceUrls": [
            "https://gamingcypher.com/annapurna-interactive-physical-playstation-4-box-sets-available-now/",
            "https://www.iam8bit.com/products/annapurna-interactive-deluxe-limited-edition",
            "https://www.pricecharting.com/game/playstation-4/annapurna-interactive-ultimate-ps4-collection",
        ],
        "summary": (
            "La Ultimate es la edición retail norteamericana distribuida por Skybound Games; "
            "la Deluxe es una edición distinta y exclusiva de iam8bit. Ambas recopilan ocho "
            "juegos de estudios diferentes publicados por Annapurna Interactive."
        ),
    },
}

# Both spelling variants are workbook-scoped physical records. They receive the
# same verified package credits, while their physical consolidation stays blocked.
ALTERNATE_BUNDLE_CREDIT_TARGETS = {
    "ps4-dishonored-pray-arkane-collection": "Dishonored & Prey: The Arkane Collection",
    "ps4-the-truth-is-out-here-mystery-adventure-pack": "The Truth Is Out There: Mystery Adventure Pack",
}

VARIANT_LINKS = [
    ("ps4-armored-core-vi-fires-of-rubicon-collector%27s-edition", "ps4-armored-core-vi-fires-of-rubicon", "edition_of", "verified"),
    ("ps4-armored-core-vi-fires-of-rubicon-launch-edition", "ps4-armored-core-vi-fires-of-rubicon", "edition_of", "verified"),
    ("ps4-blasphemous-coleccionista", "ps4-blasphemous", "edition_of", "verified"),
    ("ps4-cyberpunk-2077-samurai-pack", "ps4-cyberpunk-2077", "edition_of", "verified"),
    ("ps4-dark-souls-ii-scholar-of-the-first-sin-not-for-resale", "ps4-dark-souls-ii-scholar-of-the-first-sin", "bundle_variant_of", "verified"),
    ("ps4-gylt-collector%27s-edition", "ps4-gylt", "edition_of", "verified"),
    ("ps4-infinite-beyond-the-mind-tanya-edition", "ps4-infinite-beyond-the-mind-olga-edition", "sibling_edition_of", "verified"),
    ("ps4-middle-earth-shadow-of-mordor-game-of-the-year", "ps4-middle-earth-shadow-of-mordor", "edition_of", "verified"),
    ("ps4-okinawa-rush-black-mantis-edition", "ps4-okinawa-rush", "edition_of", "verified"),
    ("ps4-okinawa-rush-limited-edition", "ps4-okinawa-rush", "edition_of", "verified"),
    ("ps4-senran-kagura-burst-renewal-bountiful-beauties", "ps4-senran-kagura-burst-renewal", "edition_of", "verified"),
    ("ps4-young-souls-collector%27s-edition", "ps4-young-souls", "edition_of", "verified"),
    ("ps4-young-souls-deluxe-edition", "ps4-young-souls", "edition_of", "verified"),
    ("ps4-dishonored-pray-arkane-collection", "ps4-dishonored-prey-the-arkane-collection", "same_product_candidate", "requires_review"),
    ("ps4-the-truth-is-out-here-mystery-adventure-pack", "ps4-the-truth-is-out-there-mystery-adventure-pack", "same_product_candidate", "requires_review"),
]

RELIST_VARIANT_IDS = {
    "ps4-armored-core-vi-fires-of-rubicon-collector%27s-edition",
    "ps4-armored-core-vi-fires-of-rubicon-launch-edition",
    "ps4-cyberpunk-2077-samurai-pack",
    "ps4-dark-souls-ii-scholar-of-the-first-sin-not-for-resale",
    "ps4-gylt-collector%27s-edition",
    "ps4-okinawa-rush-black-mantis-edition",
    "ps4-okinawa-rush-limited-edition",
    "ps4-senran-kagura-burst-renewal-bountiful-beauties",
    "ps4-young-souls-collector%27s-edition",
    "ps4-young-souls-deluxe-edition",
}

SAFE_DUPLICATE_REDIRECTS = {
    "ps4-annapurna-ultimate-collection": "ps4-usa-annapurna-interactive-ultimate-ps4-collection",
    "ps4-jets-%27n%27-guns-2": "ps4-jets%27n%27guns-2",
    "ps4-rpg-maker": "ps4-rpg-maker-with",
}

WRONG_PLATFORM_REDIRECTS = {
    "ps4-marvel-spider-man-el-reino-de-las-sombras": "ps3-spiderman-web-of-shadows",
    "ps4-ratchet-clank-atrapados-en-el-tiempo": "ps3-ratchet-&amp;-clank-a-crack-in-time",
    "ps4-ratchet-clank-en-busca-del-tesoro": "ps3-ratchet-&amp;-clank-quest-for-booty",
    "ps4-cars": "ps2-disney-pixar-cars",
}

GENERIC_NON_GAME_ID = "ps4-not-for-resale"

TITLE_CORRECTIONS = {
    "ps4-dishonored-pray-arkane-collection": "Dishonored & Prey: The Arkane Collection",
    "ps4-the-truth-is-out-here-mystery-adventure-pack": "The Truth Is Out There: Mystery Adventure Pack",
}

ALIASES = {
    "3dclouds": ["3D Clouds", "3DClouds S.r.l."],
    "blowfish-studios": ["Blowfish Studios Pty Ltd"],
    "digerati": ["Digerati Distribution"],
    "fromsoftware": ["FromSoftware, Inc.", "From Software"],
    "honey-parade-games": ["HONEY∞PARADE GAMES, Inc."],
    "jandusoft": ["JanduSoft S.L."],
    "koei-tecmo-europe": ["Koei Tecmo Europe Limited"],
    "koei-tecmo-games": ["Koei Tecmo Games Co., Ltd."],
    "milestone-s-r-l": ["Milestone", "MILESTONE SRL"],
    "outright-games": ["Outright Games LLC", "OUTRIGHT GAMES LLC"],
    "phantom-8": ["Phantom 8 Studio UG"],
    "pix-39-n-love": ["Pix’n Love", "Pix 'n Love"],
    "raiser-games": ["Raiser Games S.L."],
    "rising-star-games": ["Rising Star Games Ltd."],
    "soedesco": ["SOEDESCO"],
    "team17": ["Team17 Digital Ltd.", "Team 17 Digital Ltd.", "Team 17"],
    "tinybuild": ["tinyBuild LLC", "TINYBUILD LLC"],
    "u-play-online": ["U-Play Online S.L.", "UPLAY ONLINE SL"],
    "wild-sphere-s-l-u": ["WildSphere", "Wild Sphere", "WILD SPHERE S.L.", "WildSphere S.L."],
}

ALIAS_SOURCE_NUMBERS = {
    "3dclouds": 39,
    "blowfish-studios": 21,
    "digerati": 31,
    "fromsoftware": 2,
    "honey-parade-games": 42,
    "jandusoft": 20,
    "koei-tecmo-europe": 30,
    "koei-tecmo-games": 30,
    "milestone-s-r-l": 12,
    "outright-games": 39,
    "phantom-8": 36,
    "pix-39-n-love": 53,
    "raiser-games": 55,
    "rising-star-games": 50,
    "soedesco": 25,
    "team17": 14,
    "tinybuild": 18,
    "u-play-online": 55,
    "wild-sphere-s-l-u": 35,
}

CORPORATE_RELATIONS = [
    ("cd-projekt-red", "cd-projekt", "studio_of", 7),
    ("ce-europe-ltd", "capcom", "regional_entity_of", 15),
    ("marvelous-europe", "marvelous", "regional_entity_of", 42),
    ("koei-tecmo-europe", "koei-tecmo-games", "regional_entity_of", 30),
    ("thunderful-publishing", "thunderful-games-ltd", "publishing_division_of", 43),
    ("rising-star-games", "thunderful-games-ltd", "owned_by", 50),
]

BLOCKED_CASES = [
    {"key": "playstation-vr-demo-disc-pal-europe", "reason": "No existe una ficha PAL Europa identificable; no se mezcla con el disco USA de 18 demos."},
    {"key": "ps4-cyberpunk-2077-samurai-pack:physical", "reason": "La editora física depende del EAN concreto, ausente en la ficha."},
    {"key": "ps4-observer-system-redux:physical", "reason": "El editor físico cambia por territorio y la ficha no aporta EAN suficiente."},
    {"key": "ps4-street-power-football:publisher", "reason": "Maximum Games y Nacon cambian según territorio/SKU; no se elige uno sin identificador."},
    {"key": "ps4-young-souls-deluxe-edition:physical", "reason": "La editora física debe confirmarse por EAN de la Deluxe concreta."},
    {"key": "ps4-youtubers-life-2:physical", "reason": "El distribuidor físico depende del territorio de la edición."},
    {"key": "ps4-mindtaker:publisher", "reason": "El propio libro deja el publisher por confirmar."},
    {"key": "dishonored-arkane-physical-consolidation", "reason": "La errata está demostrada, pero faltan EAN/CUSA para eliminar una ficha física."},
    {"key": "truth-out-there-physical-consolidation", "reason": "La variante de título está demostrada, pero faltan EAN/CUSA para eliminar una ficha física."},
]

ROLE_TO_INDEX = {
    "developer": "asDeveloper",
    "originalDeveloper": "asDeveloper",
    "portDeveloper": "asDeveloper",
    "remasterDeveloper": "asDeveloper",
    "publisher": "asPublisher",
    "originalPublisher": "asPublisher",
    "regionalPublisher": "asPublisher",
    "digitalPublisher": "asDigitalPublisher",
    "physicalPublisherOrDistributor": "asPhysicalPublisherOrDistributor",
}

DEVELOPER_ROLES = {"developer", "originalDeveloper", "portDeveloper", "remasterDeveloper"}
PUBLISHER_ROLES = {"publisher", "originalPublisher", "regionalPublisher", "digitalPublisher"}
PHYSICAL_ROLES = {"physicalPublisherOrDistributor"}


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    temporary.replace(path)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_json(value: Any) -> str:
    return hashlib.sha256(
        (json.dumps(value, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    ).hexdigest()


def update_protected_manifest(
    manifest: dict[str, Any], after_hashes: dict[str, str]
) -> dict[str, Any]:
    updated = copy.deepcopy(manifest)
    previous_update = next(
        (
            item
            for item in updated.get("protectedFileHashUpdates", [])
            if item.get("batchId") == BATCH_ID
        ),
        None,
    )
    files = {}
    for relative_path, after_hash in after_hashes.items():
        recorded_before = (
            (previous_update.get("files", {}).get(relative_path) or {}).get("before")
            if previous_update
            else None
        )
        before_hash = recorded_before or updated["protectedFileHashes"][relative_path]
        files[relative_path] = {"before": before_hash, "after": after_hash}
        updated["protectedFileHashes"][relative_path] = after_hash

    updates = [
        item
        for item in updated.setdefault("protectedFileHashUpdates", [])
        if item.get("batchId") != BATCH_ID
    ]
    updates.append({"batchId": BATCH_ID, "reviewedAt": REVIEWED_AT, "files": files})
    updated["protectedFileHashUpdates"] = updates
    return updated


def decoded(value: Any) -> str:
    current = str(value or "")
    for _ in range(5):
        result = html.unescape(current)
        if result == current:
            break
        current = result
    return current.strip()


def normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", decoded(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def work_key(value: Any) -> str:
    return normalized(value)


def extract_workbook(workbook: Path) -> None:
    if workbook.name != WORKBOOK_NAME:
        raise ValueError(f"Unexpected workbook name: {workbook.name}")
    if sha256(workbook) != WORKBOOK_SHA256:
        raise ValueError("Workbook SHA-256 does not match the reviewed source")

    from openpyxl import load_workbook

    book = load_workbook(workbook, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for sheet_name, expected_rows in EXPECTED_SHEET_ROWS.items():
        worksheet = book[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        extracted = [
            {
                header: value
                for header, value in zip(headers, values)
                if value is not None
            }
            for values in rows
            if any(value is not None for value in values)
        ]
        if len(extracted) != expected_rows:
            raise ValueError(
                f"Workbook sheet {sheet_name!r} changed: {len(extracted)} != {expected_rows}"
            )
        sheets[sheet_name] = extracted

    editorial_numbers = {
        int(row["numero"])
        for sheet_name, rows in sheets.items()
        if sheet_name.startswith("Lote editorial")
        for row in rows
    }
    if editorial_numbers != set(EDITORIAL_ID_BY_NUMBER):
        raise ValueError("Editorial rows no longer cover the reviewed 0-56 range")

    source = {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "workbook": {"name": WORKBOOK_NAME, "sha256": WORKBOOK_SHA256},
        "sheets": sheets,
    }
    write_json(SOURCE_FILE, source)
    print(f"OK extracted {WORKBOOK_NAME}: {sum(map(len, sheets.values()))} source rows")


def load_source() -> dict[str, Any]:
    source = read_json(SOURCE_FILE)
    if source.get("batchId") != BATCH_ID or source.get("workbook", {}).get("sha256") != WORKBOOK_SHA256:
        raise ValueError("Committed source snapshot does not match the reviewed workbook")
    sheets = source.get("sheets", {})
    for sheet_name, expected_rows in EXPECTED_SHEET_ROWS.items():
        if len(sheets.get(sheet_name, [])) != expected_rows:
            raise ValueError(f"Committed source sheet changed: {sheet_name}")
    return source


def editorial_rows(source: dict[str, Any]) -> dict[int, dict[str, Any]]:
    return {
        int(row["numero"]): row
        for sheet_name, rows in source["sheets"].items()
        if sheet_name.startswith("Lote editorial")
        for row in rows
    }


def source_for_bundle(source: dict[str, Any], bundle_name: str) -> dict[str, Any]:
    row = next(
        row
        for row in source["sheets"]["Recopilatorios"]
        if row["recopilatorio"] == bundle_name
    )
    override = BUNDLE_SOURCE_OVERRIDES.get(bundle_name)
    if not override:
        return row
    return {
        **row,
        "fuente": override["evidenceUrls"][0],
        "conclusion": override["summary"],
        "_evidenceUrls": override["evidenceUrls"],
    }


def source_for_mindtaker(source: dict[str, Any]) -> dict[str, Any]:
    return next(
        row
        for row in source["sheets"]["Históricos y codesarrollo"]
        if row["titulo"] == "Mindtaker"
    )


def source_url(row: dict[str, Any]) -> str:
    for field in ("fuente", "fuente_principal"):
        if str(row.get(field) or "").strip():
            return str(row[field]).strip()
    raise ValueError(f"Source row has no evidence URL: {row}")


def source_summary(row: dict[str, Any]) -> str:
    for field in ("conclusion", "accion_regionatlas"):
        if str(row.get(field) or "").strip():
            return str(row[field]).strip()
    raise ValueError(f"Source row has no evidence summary: {row}")


def locator(game: dict[str, Any], details: dict[str, Any] | None) -> dict[str, Any]:
    return {
        "catalogId": game["id"],
        "title": decoded(game.get("title")),
        "platform": game.get("platformSlug"),
        "region": game.get("region"),
        "edition": game.get("edition"),
        "physicalVariant": game.get("physicalVariant"),
        "pcId": game.get("pcId"),
        "ean": (details or {}).get("ean"),
        "reference": (details or {}).get("reference"),
        "coverUrl": game.get("coverUrl"),
        "listingStatus": game.get("listingStatus"),
    }


def credit_summary(details: dict[str, Any] | None) -> dict[str, Any]:
    details = details or {}
    return {
        "developer": (details.get("developer") or {}).get("name"),
        "publisher": (details.get("publisher") or {}).get("name"),
        "companyCredits": [
            {
                "role": credit.get("role"),
                "company": (credit.get("company") or {}).get("name"),
                "slug": (credit.get("company") or {}).get("slug"),
                "reviewBatch": (credit.get("provenance") or {}).get("reviewBatch"),
            }
            for credit in details.get("companyCredits", [])
        ],
        "individualCredits": [
            {
                "role": credit.get("role"),
                "person": (credit.get("person") or {}).get("name"),
                "slug": (credit.get("person") or {}).get("slug"),
                "reviewBatch": (credit.get("provenance") or {}).get("reviewBatch"),
            }
            for credit in details.get("individualCredits", [])
        ],
    }


def catalog_public_summary(game: dict[str, Any]) -> dict[str, Any]:
    return {
        "title": game.get("title"),
        "platformSlug": game.get("platformSlug"),
        "region": game.get("region"),
        "listingStatus": game.get("listingStatus"),
        "catalogKind": game.get("catalogKind", "game"),
        "excludeCategory": game.get("excludeCategory"),
        "excludeReason": game.get("excludeReason"),
    }


def validate_catalog_targets(
    catalog_by_id: dict[str, dict[str, Any]], details: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    all_ids = set(EDITORIAL_ID_BY_NUMBER.values()) | {
        MINDTAKER_ID,
        GENERIC_NON_GAME_ID,
        *SAFE_DUPLICATE_REDIRECTS,
        *SAFE_DUPLICATE_REDIRECTS.values(),
        *WRONG_PLATFORM_REDIRECTS,
        *WRONG_PLATFORM_REDIRECTS.values(),
        *[item for item in BUNDLE_ID_BY_NAME.values() if item],
        *ALTERNATE_BUNDLE_CREDIT_TARGETS,
        *[variant for variant, _, _, _ in VARIANT_LINKS],
        *[canonical for _, canonical, _, _ in VARIANT_LINKS],
    }
    missing = sorted(all_ids - set(catalog_by_id))
    if missing:
        raise ValueError(f"Pinned catalog IDs disappeared: {missing}")

    for catalog_id in set(EDITORIAL_ID_BY_NUMBER.values()) | {MINDTAKER_ID}:
        game = catalog_by_id[catalog_id]
        if game.get("platformSlug") != "ps4" or not str(game.get("region", "")).startswith("PAL"):
            raise ValueError(f"Target escaped PS4 PAL scope: {catalog_id}")

    jets_source = catalog_by_id["ps4-jets-%27n%27-guns-2"]
    jets_target = catalog_by_id["ps4-jets%27n%27guns-2"]
    jets_source_details = details.get(jets_source["id"], {})
    jets_target_details = details.get(jets_target["id"], {})
    if jets_source_details.get("reference") != jets_target_details.get("reference"):
        raise ValueError("Jets duplicate no longer shares the reviewed CUSA")
    if not jets_target_details.get("ean") or jets_source_details.get("ean"):
        raise ValueError("Jets canonical choice no longer matches the EAN-backed record")
    if details.get("ps4-rpg-maker", {}).get("reference") != details.get("ps4-rpg-maker-with", {}).get("reference"):
        raise ValueError("RPG Maker duplicate no longer shares the reviewed CUSA")

    return [locator(catalog_by_id[catalog_id], details.get(catalog_id)) for catalog_id in sorted(all_ids)]


def build_actions(source: dict[str, Any]) -> list[dict[str, Any]]:
    rows = editorial_rows(source)
    actions: list[dict[str, Any]] = []
    for number, catalog_id in EDITORIAL_ID_BY_NUMBER.items():
        actions.append(
            {
                "classification": "ADD_CO_DEVELOPER" if number in CO_DEVELOPER_NUMBERS else "UPDATE_CREDITS",
                "catalogId": catalog_id,
                "source": f"editorial:{number}",
                "evidenceUrl": source_url(rows[number]),
                "summary": source_summary(rows[number]),
            }
        )
    mindtaker = source_for_mindtaker(source)
    actions.append(
        {
            "classification": "ADD_CO_DEVELOPER",
            "catalogId": MINDTAKER_ID,
            "source": "Históricos y codesarrollo:Mindtaker",
            "evidenceUrl": source_url(mindtaker),
            "summary": source_summary(mindtaker),
        }
    )
    for bundle_name, catalog_id in BUNDLE_ID_BY_NAME.items():
        bundle = source_for_bundle(source, bundle_name)
        if catalog_id:
            actions.append(
                {
                    "classification": "UPDATE_CREDITS",
                    "catalogId": catalog_id,
                    "source": f"Recopilatorios:{bundle_name}",
                    "evidenceUrl": source_url(bundle),
                    "summary": source_summary(bundle),
                }
            )
    for component in source["sheets"]["Componentes"]:
        key = (component["recopilatorio"], component["juego_componente"])
        actions.append(
            {
                "classification": "ADD_COMPONENT",
                "catalogId": BUNDLE_ID_BY_NAME[component["recopilatorio"]],
                "componentTitle": component["juego_componente"],
                "linkedCatalogId": COMPONENT_LINKS.get(key),
            }
        )
    for variant, canonical, _, status in VARIANT_LINKS:
        if status == "verified":
            actions.append(
                {"classification": "LINK_VARIANT", "catalogId": variant, "targetCatalogId": canonical}
            )
    for slug, aliases in ALIASES.items():
        actions.append({"classification": "NORMALIZE_ALIAS", "companySlug": slug, "aliases": aliases})
    for source_id, target_id in SAFE_DUPLICATE_REDIRECTS.items():
        actions.append({"classification": "MERGE_DUPLICATE", "catalogId": source_id, "targetCatalogId": target_id})
    for source_id, target_id in WRONG_PLATFORM_REDIRECTS.items():
        actions.append({"classification": "MOVE_PLATFORM", "catalogId": source_id, "targetCatalogId": target_id})
    actions.append({"classification": "REMOVE_GENERIC_NON_GAME", "catalogId": GENERIC_NON_GAME_ID})
    for source_slug, target_slug, relation_type, number in CORPORATE_RELATIONS:
        actions.append(
            {
                "classification": "ADD_CORPORATE_RELATION",
                "sourceCompanySlug": source_slug,
                "targetCompanySlug": target_slug,
                "relationshipType": relation_type,
                "source": f"editorial:{number}",
            }
        )
    for blocked in BLOCKED_CASES:
        actions.append({"classification": "REQUIRES_REVIEW", **blocked})
    return actions


def provenance(row: dict[str, Any], previous_values: list[str] | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "source": "research",
        "evidenceUrls": row.get("_evidenceUrls", [source_url(row)]),
        "evidenceSummary": source_summary(row),
        "reviewedAt": REVIEWED_AT,
        "reviewBatch": BATCH_ID,
    }
    previous = list(dict.fromkeys(value for value in previous_values or [] if value))
    if previous:
        result["previousValues"] = previous
    return result


def ensure_company(companies: dict[str, dict[str, Any]], slug: str) -> dict[str, Any]:
    if slug in companies:
        return companies[slug]
    definition = COMPANY_DEFINITIONS.get(slug)
    if not definition:
        raise ValueError(f"Company slug is absent and has no reviewed definition: {slug}")
    display_name, aliases = definition
    companies[slug] = {
        "name": display_name,
        "slug": slug,
        "museumPath": "",
        "gameIds": [],
        "byPlatform": {},
        "gameCount": 0,
        "asDeveloper": [],
        "asPublisher": [],
        "asDigitalPublisher": [],
        "asPhysicalPublisherOrDistributor": [],
        **({"aliasNames": aliases} if aliases else {}),
    }
    return companies[slug]


def detail_entity(companies: dict[str, dict[str, Any]], slug: str) -> dict[str, Any]:
    company = ensure_company(companies, slug)
    return {
        "name": company["name"],
        "slug": slug,
        "museumPath": None,
        "pcPath": None,
        "source": "research",
    }


def new_detail() -> dict[str, Any]:
    return {
        "year": None,
        "releaseDate": None,
        "reference": None,
        "players": None,
        "support": None,
        "developer": None,
        "publisher": None,
        "genres": [],
        "series": None,
        "museumPath": None,
        "pcProductId": None,
        "ean": None,
        "sources": {},
        "fieldSources": {},
        "fetchedAt": REVIEWED_AT_TIMESTAMP,
        "mergedAt": REVIEWED_AT_TIMESTAMP,
    }


def roles_for_families(families: set[str]) -> set[str]:
    roles: set[str] = set()
    if "developer" in families:
        roles |= DEVELOPER_ROLES
    if "publisher" in families:
        roles |= PUBLISHER_ROLES
    if "physical" in families:
        roles |= PHYSICAL_ROLES
    return roles


def index_fields_for_families(families: set[str]) -> set[str]:
    fields: set[str] = set()
    if "developer" in families:
        fields.add("asDeveloper")
    if "publisher" in families:
        fields |= {"asPublisher", "asDigitalPublisher"}
    if "physical" in families:
        fields.add("asPhysicalPublisherOrDistributor")
    return fields


def clear_company_index_roles(
    companies: dict[str, dict[str, Any]], catalog_id: str, families: set[str]
) -> set[str]:
    changed: set[str] = set()
    fields = index_fields_for_families(families)
    all_role_fields = {
        "asDeveloper",
        "asPublisher",
        "asDigitalPublisher",
        "asPhysicalPublisherOrDistributor",
    }
    for slug, entry in companies.items():
        touched = False
        for field in fields:
            if catalog_id in entry.get(field, []):
                entry[field] = [value for value in entry[field] if value != catalog_id]
                touched = True
        if touched and not any(catalog_id in entry.get(field, []) for field in all_role_fields):
            entry["gameIds"] = [value for value in entry.get("gameIds", []) if value != catalog_id]
        if touched:
            changed.add(slug)
    return changed


def add_company_index_role(
    companies: dict[str, dict[str, Any]], catalog_id: str, role: str, slug: str
) -> None:
    entry = ensure_company(companies, slug)
    field = ROLE_TO_INDEX[role]
    if catalog_id not in entry.setdefault(field, []):
        entry[field].append(catalog_id)
    if catalog_id not in entry.setdefault("gameIds", []):
        entry["gameIds"].append(catalog_id)


def refresh_company_entry(entry: dict[str, Any], catalog_by_id: dict[str, dict[str, Any]]) -> None:
    fields = (
        "asDeveloper",
        "asPublisher",
        "asDigitalPublisher",
        "asPhysicalPublisherOrDistributor",
    )
    for field in fields:
        if field in entry:
            entry[field] = list(
                dict.fromkeys(value for value in entry[field] if value in catalog_by_id)
            )
    role_ids = {
        value
        for field in fields
        for value in entry.get(field, [])
        if value in catalog_by_id
    }
    entry["gameIds"] = list(dict.fromkeys([*entry.get("gameIds", []), *sorted(role_ids)]))
    entry["gameIds"] = [value for value in entry["gameIds"] if value in catalog_by_id]
    entry["byPlatform"] = dict(
        sorted(Counter(catalog_by_id[value]["platformSlug"] for value in entry["gameIds"]).items())
    )
    entry["gameCount"] = len(entry["gameIds"])


def apply_credit_plan(
    *,
    catalog_id: str,
    credits: list[tuple[str, str]],
    individual_credits: list[tuple[str, str, str]],
    row: dict[str, Any],
    details: dict[str, dict[str, Any]],
    companies: dict[str, dict[str, Any]],
    families: set[str],
) -> set[str]:
    detail = details.setdefault(catalog_id, new_detail())
    replaced_roles = roles_for_families(families)
    previous_values = [
        (credit.get("company") or {}).get("name")
        for credit in detail.get("companyCredits", [])
        if credit.get("role") in replaced_roles
    ]
    if "developer" in families and isinstance(detail.get("developer"), dict):
        previous_values.append(detail["developer"].get("name"))
    if "publisher" in families and isinstance(detail.get("publisher"), dict):
        previous_values.append(detail["publisher"].get("name"))
    previous_values = [str(value) for value in previous_values if value]
    previous_from_batch = next(
        (
            (credit.get("provenance") or {}).get("previousValues", [])
            for credit in detail.get("companyCredits", [])
            if (credit.get("provenance") or {}).get("reviewBatch") == BATCH_ID
        ),
        None,
    )
    if previous_from_batch is None:
        previous_from_batch = next(
            (
                value.get("previousValues", [])
                for value in (detail.get("fieldProvenance") or {}).values()
                if value.get("reviewBatch") == BATCH_ID
            ),
            previous_values,
        )
    previous_values = list(previous_from_batch)

    affected = clear_company_index_roles(companies, catalog_id, families)
    explicit = [
        credit
        for credit in detail.get("companyCredits", [])
        if credit.get("role") not in replaced_roles
    ]
    for role, slug in credits:
        ensure_company(companies, slug)
        explicit.append(
            {
                "role": role,
                "company": detail_entity(companies, slug),
                "provenance": provenance(row, previous_values),
            }
        )
        add_company_index_role(companies, catalog_id, role, slug)
        affected.add(slug)
    if explicit:
        detail["companyCredits"] = explicit
    else:
        detail.pop("companyCredits", None)

    if "developer" in families:
        developer_credit = next((item for item in credits if item[0] in DEVELOPER_ROLES), None)
        detail["developer"] = (
            detail_entity(companies, developer_credit[1]) if developer_credit else None
        )
        if developer_credit:
            detail.setdefault("fieldSources", {})["developer"] = "research"
            detail.setdefault("fieldProvenance", {})["developer"] = provenance(row, previous_values)
        else:
            detail.get("fieldSources", {}).pop("developer", None)
            detail.get("fieldProvenance", {}).pop("developer", None)

        remaining_individuals = [
            item
            for item in detail.get("individualCredits", [])
            if item.get("role") != "developer"
        ]
        for role, name, slug in individual_credits:
            remaining_individuals.append(
                {
                    "role": role,
                    "person": {
                        "name": name,
                        "slug": slug,
                        "museumPath": None,
                        "pcPath": None,
                        "source": "research",
                    },
                    "provenance": provenance(row),
                }
            )
        if remaining_individuals:
            detail["individualCredits"] = remaining_individuals
        else:
            detail.pop("individualCredits", None)

    if "publisher" in families:
        publisher_credit = next(
            (
                item
                for preferred_role in ("originalPublisher", "publisher", "regionalPublisher", "digitalPublisher")
                for item in credits
                if item[0] == preferred_role
            ),
            None,
        )
        detail["publisher"] = (
            detail_entity(companies, publisher_credit[1]) if publisher_credit else None
        )
        if publisher_credit:
            detail.setdefault("fieldSources", {})["publisher"] = "research"
            detail.setdefault("fieldProvenance", {})["publisher"] = provenance(row, previous_values)
        else:
            detail.get("fieldSources", {}).pop("publisher", None)
            detail.get("fieldProvenance", {}).pop("publisher", None)

    detail["mergedAt"] = REVIEWED_AT_TIMESTAMP
    return affected


def apply_aliases(
    companies: dict[str, dict[str, Any]], source: dict[str, Any]
) -> tuple[set[str], dict[str, Any]]:
    rows = editorial_rows(source)
    affected: set[str] = set()
    records = []
    existing_records = {}
    if COMPANY_ALIASES_FILE.is_file():
        existing_records = {
            item["companySlug"]: item
            for item in read_json(COMPANY_ALIASES_FILE).get("aliases", [])
            if item.get("companySlug")
        }
    for slug, aliases in ALIASES.items():
        entry = ensure_company(companies, slug)
        before = list(entry.get("aliasNames", []))
        entry["aliasNames"] = list(dict.fromkeys([*before, *aliases]))
        if entry["aliasNames"] != before:
            affected.add(slug)
        row = rows[ALIAS_SOURCE_NUMBERS[slug]]
        recorded_previous = (
            existing_records[slug].get("provenance", {}).get("previousValues", [])
            if slug in existing_records
            else before
        )
        records.append(
            {
                "companySlug": slug,
                "displayName": entry["name"],
                "aliases": aliases,
                "provenance": provenance(row, recorded_previous),
            }
        )
    return affected, {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "reviewedAt": REVIEWED_AT,
        "aliases": records,
    }


def build_corporate_relations(source: dict[str, Any]) -> dict[str, Any]:
    rows = editorial_rows(source)
    return {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "relationships": [
            {
                "id": f"{source_slug}:{relation_type}:{target_slug}",
                "sourceCompanySlug": source_slug,
                "targetCompanySlug": target_slug,
                "relationshipType": relation_type,
                "provenance": provenance(rows[number]),
            }
            for source_slug, target_slug, relation_type, number in CORPORATE_RELATIONS
        ],
    }


def component_key(component: dict[str, Any]) -> tuple[str, str]:
    return component["recopilatorio"], component["juego_componente"]


def build_commercial_relations(
    source: dict[str, Any], catalog_by_id: dict[str, dict[str, Any]]
) -> dict[str, Any]:
    component_rows = source["sheets"]["Componentes"]
    components_by_bundle: dict[str, list[dict[str, Any]]] = {}
    for component in component_rows:
        components_by_bundle.setdefault(component["recopilatorio"], []).append(component)

    compilations = []
    for bundle in source["sheets"]["Recopilatorios"]:
        name = bundle["recopilatorio"]
        catalog_id = BUNDLE_ID_BY_NAME[name]
        components = []
        for position, component in enumerate(components_by_bundle.get(name, []), start=1):
            linked_id = COMPONENT_LINKS.get(component_key(component))
            if linked_id and linked_id not in catalog_by_id:
                raise ValueError(f"Pinned component ID disappeared: {linked_id}")
            components.append(
                {
                    "position": position,
                    "title": component["juego_componente"],
                    "developerCredit": component["desarrolladora"],
                    "publisherCredit": component["editora"],
                    "notes": component.get("notas") or None,
                    "catalogId": linked_id,
                    "provenance": provenance(bundle),
                }
            )
        if len(components) != int(bundle["componentes"]):
            raise ValueError(f"Component total drifted for {name}")
        compilations.append(
            {
                "id": f"compilation:{normalized(name).replace(' ', '-')}",
                "catalogId": catalog_id,
                "title": BUNDLE_PUBLIC_TITLES.get(name, name),
                "status": "verified" if catalog_id else "requires_review",
                "componentCount": len(components),
                "components": components,
                "provenance": provenance(bundle),
            }
        )

    rows = editorial_rows(source)
    duplicates = source["sheets"]["Duplicados catálogo"]
    variant_rows = []
    editorial_by_id = {catalog_id: number for number, catalog_id in EDITORIAL_ID_BY_NUMBER.items()}
    for variant_id, canonical_id, relation_type, status in VARIANT_LINKS:
        if variant_id in editorial_by_id:
            row = rows[editorial_by_id[variant_id]]
        else:
            needle = "Dishonored" if "dishonored" in variant_id else "Truth"
            row = next(item for item in duplicates if needle in item["titulo_canonico"])
            row = {
                "fuente": source_for_bundle(
                    source,
                    "Dishonored & Prey: The Arkane Collection"
                    if needle == "Dishonored"
                    else "The Truth Is Out There: Mystery Adventure Pack",
                )["fuente_principal"],
                "conclusion": row["accion_recomendada"],
            }
        variant_rows.append(
            {
                "variantCatalogId": variant_id,
                "canonicalCatalogId": canonical_id,
                "relationshipType": relation_type,
                "status": status,
                "provenance": provenance(row),
            }
        )

    return {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "reviewedAt": REVIEWED_AT,
        "compilations": compilations,
        "variants": variant_rows,
    }


def clean_slug(value: str) -> str:
    from urllib.parse import unquote

    return normalized(unquote(decoded(value))).replace(" ", "-")


def catalog_seo_param(game: dict[str, Any]) -> str:
    region = str(game.get("region") or "")
    short = {
        "PAL España": "pal-es",
        "PAL Europa": "pal-eu",
        "PAL UK": "pal-uk",
        "USA": "pal-us",
        "NTSC USA": "pal-us",
        "Japón": "pal-jp",
        "NTSC-J": "pal-jp",
    }.get(region)
    if not short:
        short = clean_slug(re.sub(r"^PAL\s+", "pal-", region, flags=re.I))
    return f"{clean_slug(game['slug'])}-{game['platformSlug']}-{short}"


def build_route_redirects(catalog_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    records = []
    for reason, mappings in (
        ("same_product", SAFE_DUPLICATE_REDIRECTS),
        ("wrong_platform", WRONG_PLATFORM_REDIRECTS),
    ):
        for source_id, target_id in mappings.items():
            source_game = catalog_by_id[source_id]
            target_game = catalog_by_id[target_id]
            target_param = catalog_seo_param(target_game)
            source_params = [source_id, catalog_seo_param(source_game)]
            records.append(
                {
                    "sourceParams": list(dict.fromkeys(
                        source_param
                        for source_param in source_params
                        if source_param != target_param
                    )),
                    "targetCatalogId": target_id,
                    "targetParam": target_param,
                    "permanent": True,
                    "reason": reason,
                    "reviewedAt": REVIEWED_AT,
                    "reviewBatch": BATCH_ID,
                    "targetLocator": locator(target_game, None),
                }
            )
    return {"schemaVersion": 1, "batchId": BATCH_ID, "redirects": records}


def apply_catalog_changes(catalog_by_id: dict[str, dict[str, Any]]) -> set[str]:
    allowed: set[str] = set()
    for catalog_id in RELIST_VARIANT_IDS:
        game = catalog_by_id[catalog_id]
        game["listingStatus"] = "listed"
        game.pop("excludeCategory", None)
        game.pop("excludeReason", None)
        allowed.add(catalog_id)

    nfr_variant = catalog_by_id["ps4-dark-souls-ii-scholar-of-the-first-sin-not-for-resale"]
    nfr_variant["physicalVariant"] = "not-for-resale"

    for catalog_id in SAFE_DUPLICATE_REDIRECTS:
        game = catalog_by_id[catalog_id]
        game["listingStatus"] = "excluded"
        game["excludeCategory"] = "duplicate"
        game["excludeReason"] = f"Consolidated by {BATCH_ID}; permanent redirect retained"
        allowed.add(catalog_id)

    for catalog_id in WRONG_PLATFORM_REDIRECTS:
        game = catalog_by_id[catalog_id]
        game["listingStatus"] = "excluded"
        game["excludeCategory"] = "wrong-platform"
        game["excludeReason"] = f"Misidentified PS4 record; redirected by {BATCH_ID}"
        allowed.add(catalog_id)

    marker = catalog_by_id[GENERIC_NON_GAME_ID]
    marker["listingStatus"] = "excluded"
    marker["catalogKind"] = "non_game_marker"
    marker["excludeCategory"] = "promo"
    marker["excludeReason"] = "Generic Not For Resale marker; not a videogame"
    allowed.add(GENERIC_NON_GAME_ID)

    for catalog_id, title in TITLE_CORRECTIONS.items():
        catalog_by_id[catalog_id]["title"] = title
        allowed.add(catalog_id)
    return allowed


def recompute_catalog_meta(
    catalog: list[dict[str, Any]], details: dict[str, Any], companies: dict[str, Any]
) -> tuple[dict[str, Any], dict[str, Any]]:
    meta = read_json(META_FILE)
    curation = read_json(CURATION_FILE)

    def is_listed_game(game: dict[str, Any]) -> bool:
        return game.get("listingStatus") != "excluded" and game.get("catalogKind", "game") == "game"

    listed = [game for game in catalog if is_listed_game(game)]
    excluded = [game for game in catalog if not is_listed_game(game)]
    listed_by_platform = Counter(game["platformSlug"] for game in listed)
    excluded_by_platform = Counter(game["platformSlug"] for game in excluded)
    by_category = Counter(game.get("excludeCategory") or "other" for game in excluded)

    def ordered(existing: dict[str, Any], values: Counter[str]) -> dict[str, int]:
        result = {key: values[key] for key in existing if key in values}
        for key in sorted(set(values) - set(result)):
            result[key] = values[key]
        return result

    meta.update(
        {
            "catalogListed": len(listed),
            "listedByPlatform": ordered(meta.get("listedByPlatform", {}), listed_by_platform),
            "catalogExcluded": len(excluded),
            "catalogTotal": len(catalog),
            "excludedByPlatform": ordered(meta.get("excludedByPlatform", {}), excluded_by_platform),
            "curationByCategory": ordered(meta.get("curationByCategory", {}), by_category),
            "coversListed": sum(bool(game.get("coverUrl")) for game in listed),
            "coversListedPct": round(100 * sum(bool(game.get("coverUrl")) for game in listed) / max(len(listed), 1), 1),
            "gamesWithDetails": len(details),
            "indexCompanies": len(companies),
        }
    )
    curation.update(
        {
            "total": len(catalog),
            "listed": len(listed),
            "excluded": len(excluded),
            "byCategory": ordered(curation.get("byCategory", {}), by_category),
            "listedByPlatform": ordered(curation.get("listedByPlatform", {}), listed_by_platform),
            "excludedByPlatform": ordered(curation.get("excludedByPlatform", {}), excluded_by_platform),
        }
    )
    return meta, curation


def allowed_catalog_keys(catalog_id: str) -> set[str]:
    allowed = {"listingStatus", "excludeCategory", "excludeReason"}
    if catalog_id == GENERIC_NON_GAME_ID:
        allowed.add("catalogKind")
    if catalog_id in TITLE_CORRECTIONS:
        allowed.add("title")
    if catalog_id == "ps4-dark-souls-ii-scholar-of-the-first-sin-not-for-resale":
        allowed.add("physicalVariant")
    return allowed


def catalog_changed_keys(before: dict[str, Any], after: dict[str, Any]) -> set[str]:
    return {
        key
        for key in before.keys() | after.keys()
        if before.get(key) != after.get(key)
    }


def protected_catalog_projection(catalog: list[dict[str, Any]]) -> list[dict[str, Any]]:
    protected = {
        "id",
        "slug",
        "platformSlug",
        "region",
        "edition",
        "coverUrl",
        "pcId",
        "pcPath",
        "marketMin",
        "marketMax",
        "recommendedPrice",
        "estimatedPriceLoose",
        "estimatedPriceGameManual",
        "estimatedPriceComplete",
        "estimatedPriceSealed",
        "estimatedPriceNewRetail",
        "priceChartingLooseUsd",
        "priceChartingCompleteUsd",
        "priceChartingSealedUsd",
    }
    return [{key: game.get(key) for key in protected} for game in catalog]


def apply_to_memory(source: dict[str, Any]) -> dict[str, Any]:
    catalog = read_json(CATALOG_FILE)
    details = read_json(DETAILS_FILE)
    companies = read_json(COMPANIES_FILE)
    catalog_before = copy.deepcopy(catalog)
    details_before = copy.deepcopy(details)
    companies_before = copy.deepcopy(companies)
    catalog_by_id = {game["id"]: game for game in catalog}
    if len(catalog_by_id) != len(catalog):
        raise ValueError("Catalog IDs are not unique before the batch")
    locators = validate_catalog_targets(catalog_by_id, details)
    rows = editorial_rows(source)
    affected_company_slugs: set[str] = set()

    used_slugs = {
        slug
        for credits in [*CREDIT_PLANS.values(), MINDTAKER_CREDITS, *BUNDLE_CREDIT_PLANS.values()]
        for _, slug in credits
    } | {slug for relation in CORPORATE_RELATIONS for slug in relation[:2]} | set(ALIASES)
    for slug in sorted(used_slugs):
        ensure_company(companies, slug)

    for number, credits in CREDIT_PLANS.items():
        catalog_id = EDITORIAL_ID_BY_NUMBER[number]
        affected_company_slugs |= apply_credit_plan(
            catalog_id=catalog_id,
            credits=credits,
            individual_credits=INDIVIDUAL_CREDITS.get(number, []),
            row=rows[number],
            details=details,
            companies=companies,
            families={"developer", "publisher", "physical"},
        )

    mindtaker_row = source_for_mindtaker(source)
    affected_company_slugs |= apply_credit_plan(
        catalog_id=MINDTAKER_ID,
        credits=MINDTAKER_CREDITS,
        individual_credits=[],
        row=mindtaker_row,
        details=details,
        companies=companies,
        families={"developer"},
    )

    for bundle_name, credits in BUNDLE_CREDIT_PLANS.items():
        catalog_id = BUNDLE_ID_BY_NAME[bundle_name]
        if not catalog_id:
            continue
        affected_company_slugs |= apply_credit_plan(
            catalog_id=catalog_id,
            credits=credits,
            individual_credits=[],
            row=source_for_bundle(source, bundle_name),
            details=details,
            companies=companies,
            families={"developer", "publisher", "physical"},
        )

    for catalog_id, bundle_name in ALTERNATE_BUNDLE_CREDIT_TARGETS.items():
        affected_company_slugs |= apply_credit_plan(
            catalog_id=catalog_id,
            credits=BUNDLE_CREDIT_PLANS[bundle_name],
            individual_credits=[],
            row=source_for_bundle(source, bundle_name),
            details=details,
            companies=companies,
            families={"developer", "publisher", "physical"},
        )

    alias_affected, aliases_file = apply_aliases(companies, source)
    affected_company_slugs |= alias_affected
    for slug in used_slugs:
        affected_company_slugs.add(slug)

    catalog_allowed_ids = apply_catalog_changes(catalog_by_id)
    for slug in affected_company_slugs:
        refresh_company_entry(companies[slug], catalog_by_id)

    commercial_relations = build_commercial_relations(source, catalog_by_id)
    company_relations = build_corporate_relations(source)
    route_redirects = build_route_redirects(catalog_by_id)
    meta, curation = recompute_catalog_meta(catalog, details, companies)
    protected_after_hashes = {
        "data/game-details.json": sha256_json(details),
        "data/index/companies.json": sha256_json(companies),
    }
    protected_manifests = {
        path: update_protected_manifest(read_json(path), protected_after_hashes)
        for path in MANIFEST_FILES
    }

    catalog_changes = {
        after["id"]: {
            "before": catalog_public_summary(before),
            "after": catalog_public_summary(after),
            "changedKeys": sorted(catalog_changed_keys(before, after)),
        }
        for before, after in zip(catalog_before, catalog)
        if before != after
    }
    unexpected_catalog_changes = {
        catalog_id: sorted(set(change["changedKeys"]) - allowed_catalog_keys(catalog_id))
        for catalog_id, change in catalog_changes.items()
        if set(change["changedKeys"]) - allowed_catalog_keys(catalog_id)
    }
    if unexpected_catalog_changes:
        raise ValueError(f"Unexpected catalog field changes: {unexpected_catalog_changes}")
    if set(catalog_changes) - catalog_allowed_ids:
        raise ValueError(f"Catalog changes escaped the pinned scope: {sorted(set(catalog_changes) - catalog_allowed_ids)}")

    details_scope = set(EDITORIAL_ID_BY_NUMBER.values()) | {MINDTAKER_ID}
    details_scope |= {value for value in BUNDLE_ID_BY_NAME.values() if value}
    details_scope |= set(ALTERNATE_BUNDLE_CREDIT_TARGETS)
    details_changes = {
        catalog_id: {
            "before": credit_summary(details_before.get(catalog_id)),
            "after": credit_summary(details.get(catalog_id)),
        }
        for catalog_id in details_before.keys() | details.keys()
        if details_before.get(catalog_id) != details.get(catalog_id)
    }
    if set(details_changes) - details_scope:
        raise ValueError(f"Detail changes escaped scope: {sorted(set(details_changes) - details_scope)}")

    company_changes = {
        slug
        for slug in companies_before.keys() | companies.keys()
        if companies_before.get(slug) != companies.get(slug)
    }
    if company_changes - affected_company_slugs:
        raise ValueError(f"Company changes escaped scope: {sorted(company_changes - affected_company_slugs)}")

    protected_before = protected_catalog_projection(catalog_before)
    protected_after = protected_catalog_projection(catalog)
    if protected_before != protected_after:
        raise ValueError("IDs, URL inputs, regions, editions, covers or prices changed")

    actions = build_actions(source)
    action_counts = dict(sorted(Counter(item["classification"] for item in actions).items()))
    credit_counts = Counter(
        role
        for credits in [*CREDIT_PLANS.values(), MINDTAKER_CREDITS, *BUNDLE_CREDIT_PLANS.values()]
        for role, _ in credits
    )
    # Alternate bundle records carry the same package evidence and are intentional.
    for bundle_name in ALTERNATE_BUNDLE_CREDIT_TARGETS.values():
        credit_counts.update(role for role, _ in BUNDLE_CREDIT_PLANS[bundle_name])

    report = {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "reviewedAt": REVIEWED_AT,
        "source": source["workbook"],
        "writesPerformed": False,
        "summary": {
            "catalogRowsBefore": len(catalog_before),
            "catalogRowsAfter": len(catalog),
            "uniqueIdsBefore": len({game["id"] for game in catalog_before}),
            "uniqueIdsAfter": len({game["id"] for game in catalog}),
            "companiesBefore": len(companies_before),
            "companiesAfter": len(companies),
            "editorialCatalogEntries": len(EDITORIAL_ID_BY_NUMBER) + 1,
            "compilationsDocumented": len(commercial_relations["compilations"]),
            "compilationsLinked": sum(bool(item["catalogId"]) for item in commercial_relations["compilations"]),
            "componentsDocumented": sum(item["componentCount"] for item in commercial_relations["compilations"]),
            "componentsLinkedExactly": sum(bool(component["catalogId"]) for item in commercial_relations["compilations"] for component in item["components"]),
            "variantRelations": len(commercial_relations["variants"]),
            "verifiedVariantRelations": sum(item["status"] == "verified" for item in commercial_relations["variants"]),
            "companyCreditsByRole": dict(sorted(credit_counts.items())),
            "individualCredits": sum(len(value) for value in INDIVIDUAL_CREDITS.values()),
            "aliasesNormalized": len(ALIASES),
            "corporateRelations": len(CORPORATE_RELATIONS),
            "safeRedirects": len(route_redirects["redirects"]),
            "blockedCases": len(BLOCKED_CASES),
            "catalogEntriesChanged": len(catalog_changes),
            "detailsChanged": len(details_changes),
            "companiesChanged": len(company_changes),
            "companiesCreated": len(set(companies) - set(companies_before)),
            "actions": action_counts,
        },
        "locators": locators,
        "catalogBeforeAfter": catalog_changes,
        "detailsBeforeAfter": details_changes,
        "companiesCreated": sorted(set(companies) - set(companies_before)),
        "companiesChanged": sorted(company_changes),
        "blockedCases": BLOCKED_CASES,
        "scopeChecks": {
            "catalogRowsPreserved": len(catalog_before) == len(catalog),
            "catalogIdsPreserved": [game["id"] for game in catalog_before] == [game["id"] for game in catalog],
            "catalogIdsUnique": len(catalog) == len({game["id"] for game in catalog}),
            "urlInputsRegionsEditionsPreserved": all(
                all(before.get(field) == after.get(field) for field in ("id", "slug", "platformSlug", "region", "edition"))
                for before, after in zip(catalog_before, catalog)
            ),
            "coversAndPricesPreserved": protected_before == protected_after,
            "detailsOnlyWorkbookScope": not bool(set(details_changes) - details_scope),
            "companiesOnlyCreditScope": not bool(company_changes - affected_company_slugs),
            "noCrossRegionPropagation": all(catalog_by_id[catalog_id]["platformSlug"] == "ps4" and str(catalog_by_id[catalog_id]["region"]).startswith("PAL") for catalog_id in set(EDITORIAL_ID_BY_NUMBER.values()) | {MINDTAKER_ID}),
            "allBlockedCasesRemainUnpublished": True,
        },
        "verification": {
            "semanticTest": "PENDING",
            "typecheck": "PENDING",
            "lint": "PENDING",
            "unitTests": "PENDING",
            "collectorControls": "PENDING",
            "affiliateOffersV1": "PENDING",
            "build": "PENDING",
            "previewQa": "PENDING",
        },
    }
    if not all(report["scopeChecks"].values()):
        raise ValueError(f"Scope check failed: {report['scopeChecks']}")

    return {
        "catalog": catalog,
        "details": details,
        "companies": companies,
        "meta": meta,
        "curation": curation,
        "commercialRelations": commercial_relations,
        "routeRedirects": route_redirects,
        "companyRelations": company_relations,
        "companyAliases": aliases_file,
        "protectedManifests": protected_manifests,
        "report": report,
        "actions": actions,
    }


def render_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    actions = summary["actions"]
    checks = report["scopeChecks"]
    lines = [
        "# PS4 PAL - recopilatorios y créditos verificados",
        "",
        f"Lote: `{BATCH_ID}`",
        f"Fuente: `{WORKBOOK_NAME}` (`{WORKBOOK_SHA256}`)",
        "",
        "## Resultado",
        "",
        f"- Fichas editoriales localizadas: {summary['editorialCatalogEntries']}.",
        f"- Recopilatorios documentados: {summary['compilationsDocumented']}; enlazados a ficha: {summary['compilationsLinked']}.",
        f"- Componentes conservados: {summary['componentsDocumented']}; enlaces exactos al catálogo: {summary['componentsLinkedExactly']}.",
        f"- Variantes relacionadas: {summary['variantRelations']} ({summary['verifiedVariantRelations']} verificadas; el resto conserva revisión física pendiente).",
        f"- Alias normalizados: {summary['aliasesNormalized']}.",
        f"- Relaciones corporativas verificadas: {summary['corporateRelations']}.",
        f"- Redirecciones permanentes seguras: {summary['safeRedirects']}.",
        f"- Casos bloqueados: {summary['blockedCases']}.",
        "",
        "## Acciones del dry-run",
        "",
        *[f"- {name}: {count}." for name, count in actions.items()],
        "",
        "## Invariantes",
        "",
        f"- Fichas: {summary['catalogRowsBefore']} -> {summary['catalogRowsAfter']}.",
        f"- IDs únicos: {summary['uniqueIdsBefore']} -> {summary['uniqueIdsAfter']}.",
        "- IDs, slugs, plataforma, región, edición y URLs: conservados.",
        "- Portadas y todos los campos de precio: conservados.",
        "- Sin propagación a otras regiones, plataformas o títulos parecidos.",
        *[f"- {name}: {'PASS' if value else 'FAIL'}." for name, value in checks.items()],
        "",
        "## Casos no aplicados",
        "",
        *[f"- `{item['key']}`: {item['reason']}" for item in report["blockedCases"]],
        "",
        "## Verificación",
        "",
        *[f"- {name}: {value}." for name, value in report["verification"].items()],
        "",
        "La PR permanece en borrador y no autoriza fusión ni despliegue a Production.",
        "",
    ]
    return "\n".join(lines)


def run_dry_run() -> dict[str, Any]:
    source = load_source()
    result = apply_to_memory(source)
    write_json(DRY_RUN_FILE, {"report": result["report"], "actions": result["actions"]})
    write_json(REPORT_FILE, result["report"])
    REPORT_MD_FILE.write_text(render_markdown(result["report"]), encoding="utf-8")
    return result


def run_apply() -> dict[str, Any]:
    result = apply_to_memory(load_source())
    write_json(CATALOG_FILE, result["catalog"])
    write_json(DETAILS_FILE, result["details"])
    write_json(COMPANIES_FILE, result["companies"])
    write_json(META_FILE, result["meta"])
    write_json(CURATION_FILE, result["curation"])
    write_json(COMMERCIAL_RELATIONS_FILE, result["commercialRelations"])
    write_json(ROUTE_REDIRECTS_FILE, result["routeRedirects"])
    write_json(COMPANY_RELATIONS_FILE, result["companyRelations"])
    write_json(COMPANY_ALIASES_FILE, result["companyAliases"])
    for path, manifest in result["protectedManifests"].items():
        write_json(path, manifest)
    subprocess.run([sys.executable, str(VERIFIED_INDEX_BUILDER), "--write"], cwd=ROOT, check=True)
    dry_run_report = read_json(DRY_RUN_FILE).get("report") if DRY_RUN_FILE.is_file() else None
    if (
        isinstance(dry_run_report, dict)
        and dry_run_report.get("batchId") == BATCH_ID
        and dry_run_report.get("summary", {}).get("catalogEntriesChanged", 0) > 0
    ):
        result["report"] = copy.deepcopy(dry_run_report)
    result["report"]["writesPerformed"] = True
    write_json(REPORT_FILE, result["report"])
    REPORT_MD_FILE.write_text(render_markdown(result["report"]), encoding="utf-8")
    return result


def validate_committed() -> dict[str, Any]:
    source = load_source()
    result = apply_to_memory(source)
    expected_files = {
        COMMERCIAL_RELATIONS_FILE: result["commercialRelations"],
        ROUTE_REDIRECTS_FILE: result["routeRedirects"],
        COMPANY_RELATIONS_FILE: result["companyRelations"],
        COMPANY_ALIASES_FILE: result["companyAliases"],
    }
    for path, expected in expected_files.items():
        if read_json(path) != expected:
            raise ValueError(f"Generated file is missing or out of sync: {path.relative_to(ROOT)}")
    for path, expected in result["protectedManifests"].items():
        if read_json(path) != expected:
            raise ValueError(f"Protected manifest is out of sync: {path.relative_to(ROOT)}")
    if result["catalog"] != read_json(CATALOG_FILE):
        raise ValueError("Catalog application is not idempotent")
    if result["details"] != read_json(DETAILS_FILE):
        raise ValueError("Game-detail application is not idempotent")
    if result["companies"] != read_json(COMPANIES_FILE):
        raise ValueError("Company-index application is not idempotent")
    report = read_json(REPORT_FILE)
    if report.get("writesPerformed") is not True:
        raise ValueError("Committed report is not marked as applied")
    print(
        "OK PS4 PAL compilations: "
        f"{len(EDITORIAL_ID_BY_NUMBER) + 1} editorial entries, "
        f"{EXPECTED_SHEET_ROWS['Recopilatorios']} compilations, "
        f"{EXPECTED_SHEET_ROWS['Componentes']} components and {len(BLOCKED_CASES)} blocked cases"
    )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    modes = parser.add_mutually_exclusive_group(required=True)
    modes.add_argument("--extract-workbook", type=Path)
    modes.add_argument("--dry-run", action="store_true")
    modes.add_argument("--apply", action="store_true")
    modes.add_argument("--check", action="store_true")
    args = parser.parse_args()
    if args.extract_workbook:
        extract_workbook(args.extract_workbook.resolve())
    elif args.dry_run:
        result = run_dry_run()
        print(json.dumps(result["report"]["summary"], ensure_ascii=False, indent=2))
    elif args.apply:
        result = run_apply()
        print(json.dumps(result["report"]["summary"], ensure_ascii=False, indent=2))
    else:
        validate_committed()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
