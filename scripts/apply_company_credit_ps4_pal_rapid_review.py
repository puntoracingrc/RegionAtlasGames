#!/usr/bin/env python3
"""Apply the role-separated PS4 PAL rapid-review workbook with strict guards."""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import html
import json
import re
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RESEARCH = DATA / "research"
CATALOG_FILE = DATA / "catalog.json"
DETAILS_FILE = DATA / "game-details.json"
COMPANIES_FILE = DATA / "index/companies.json"
COMPANY_ENTITIES_FILE = DATA / "index/company-entities.json"
WORK_IDENTITIES_FILE = DATA / "index/catalog-work-identities.json"
META_FILE = DATA / "meta.json"
CURATION_FILE = DATA / "curation-report.json"
VERIFIED_INDEX_BUILDER = ROOT / "scripts/build_verified_company_credit_index.py"
MANIFEST_FILES = (
    RESEARCH / "company-study/manifest.json",
    RESEARCH / "person-study/manifest.json",
)

BATCH_ID = "company-credit-ps4-pal-rapid-review-2026-09-05"
REVIEWED_AT = "2026-09-05"
REVIEWED_AT_TIMESTAMP = "2026-09-05T12:00:00Z"
SUCCESSOR_BATCH_ID = "company-credit-ps4-pal-compilations-2026-09-05"
ALLOWED_SUCCESSOR_ROLE_CREDITS = {
    ("ps4-annapurna-ultimate-collection", "physicalPublisherOrDistributor"): ("skybound-games",),
    ("ps4-blasphemous-coleccionista", "developer"): ("the-game-kitchen",),
    ("ps4-crash-team-racing-&amp;-spyro-reignited-trilogy", "physicalPublisherOrDistributor"): ("activision",),
    ("ps4-cyberpunk-2077-samurai-pack", "developer"): ("cd-projekt-red",),
    ("ps4-dark-thrones-&amp;-witch-hunter-double-pack", "physicalPublisherOrDistributor"): ("funbox-media",),
    ("ps4-dishonored-pray-arkane-collection", "physicalPublisherOrDistributor"): ("bethesda-softworks",),
    ("ps4-doctor-who-duo-bundle", "physicalPublisherOrDistributor"): ("maximum-games",),
    ("ps4-middle-earth-shadow-of-mordor-game-of-the-year", "developer"): ("monolith-productions",),
    ("ps4-persona-5-&#43;-persona-5-dancing-the-starlight-day-one-edition-2-hits-pack", "physicalPublisherOrDistributor"): ("atlus",),
    ("ps4-playlink-games-collection", "physicalPublisherOrDistributor"): ("sony-interactive-entertainment",),
    ("ps4-senran-kagura-burst-renewal-bountiful-beauties", "developer"): ("honey-parade-games",),
    ("ps4-spyro-reignited-trilogy-&amp;-crash-bandicoot-n-sane-trilogy", "physicalPublisherOrDistributor"): ("activision",),
    ("ps4-terra-trilogy", "physicalPublisherOrDistributor"): ("funbox-media",),
    ("ps4-the-truth-is-out-here-mystery-adventure-pack", "physicalPublisherOrDistributor"): ("uig-entertainment",),
    ("ps4-the-truth-is-out-there-mystery-adventure-pack", "physicalPublisherOrDistributor"): ("uig-entertainment",),
    ("ps4-wolfenstein-alt-history-collection", "physicalPublisherOrDistributor"): ("bethesda-softworks",),
    ("ps4-wonder-boy-collection", "physicalPublisherOrDistributor"): ("inin-games",),
    ("ps4-world-of-simulators", "physicalPublisherOrDistributor"): ("uig-entertainment",),
}
SUCCESSOR_CATALOG_IDS = {
    "ps4-annapurna-ultimate-collection": "ps4-usa-annapurna-interactive-ultimate-ps4-collection",
}
WORKBOOK_NAME = "RegionAtlas_PS4_PAL_repaso_rapido_creditos_2026-09-05.xlsx"
WORKBOOK_SHA256 = "f8a1589ad8cfd168c202a0c6173b64994fd1c127c01ee955861b0c7226531de1"

SOURCE_RESOLVED = RESEARCH / "ps4-pal-rapid-review-resolved.csv"
SOURCE_CONFLICTS = RESEARCH / "ps4-pal-rapid-review-conflicts.csv"
SOURCE_NON_GAMES = RESEARCH / "ps4-pal-rapid-review-non-games.csv"
DECISIONS_FILE = RESEARCH / "ps4-pal-rapid-review-decisions.csv"
BLOCKED_FILE = RESEARCH / "ps4-pal-rapid-review-blocked.json"
NON_GAME_DECISIONS_FILE = RESEARCH / "ps4-pal-rapid-review-non-game-decisions.csv"
COMPANY_RESOLUTION_FILE = RESEARCH / "ps4-pal-rapid-review-company-resolution.csv"
REPORT_FILE = RESEARCH / "company-credit-ps4-pal-rapid-review-report.json"
REPORT_MD_FILE = RESEARCH / "company-credit-ps4-pal-rapid-review-report.md"

EXPECTED = {
    "sourceReportedReviewedRoles": 1760,
    "extractedRoleRows": 1750,
    "resolvedRows": 1655,
    "resolvedCatalogIds": 1123,
    "conflicts": 95,
    "conflictCatalogIds": 90,
    "nonGames": 6,
    "scopeCatalogIds": 1199,
    "sourceReportedWorkIdentities": 1071,
    "workIdentities": 1070,
}
EXPECTED_ACTIONS = {"ADD": 1499, "REPLACE": 74, "CONFIRM": 82, "SKIP_CHANGED": 0}
EXPECTED_ROLES = {
    "DEVELOPER": 656,
    "DIGITAL_PUBLISHER": 60,
    "PHYSICAL_PUBLISHER_OR_DISTRIBUTOR": 937,
    "PUBLISHER": 2,
}
ROLE_NAMES = {
    "DEVELOPER": "developer",
    "PUBLISHER": "publisher",
    "DIGITAL_PUBLISHER": "digitalPublisher",
    "PHYSICAL_PUBLISHER_OR_DISTRIBUTOR": "physicalPublisherOrDistributor",
}
ROLE_INDEX_FIELDS = {
    "DEVELOPER": "asDeveloper",
    "PUBLISHER": "asPublisher",
    "DIGITAL_PUBLISHER": "asDigitalPublisher",
    "PHYSICAL_PUBLISHER_OR_DISTRIBUTOR": "asPhysicalPublisherOrDistributor",
}
PUBLISHER_ROLES = {
    "PUBLISHER",
    "DIGITAL_PUBLISHER",
    "PHYSICAL_PUBLISHER_OR_DISTRIBUTOR",
}


def allowed_successor_role_credits(row: dict[str, str], detail: dict[str, Any]) -> list[dict[str, Any]]:
    role_name = ROLE_NAMES[row["role"]]
    allowed_slugs = ALLOWED_SUCCESSOR_ROLE_CREDITS.get((row["catalog_id"], role_name))
    if not allowed_slugs:
        return []
    credits = [
        credit
        for credit in detail.get("companyCredits", [])
        if credit.get("role") == role_name
    ]
    if tuple(credit.get("company", {}).get("slug") for credit in credits) != allowed_slugs:
        return []
    expected_previous = set(row["company_display_names"].split(" | "))
    catalog_was_reassigned = row["catalog_id"] in SUCCESSOR_CATALOG_IDS
    for credit in credits:
        provenance = credit.get("provenance", {})
        if (
            provenance.get("reviewBatch") != SUCCESSOR_BATCH_ID
            or provenance.get("reviewedAt") != REVIEWED_AT
            or not provenance.get("evidenceUrls")
            or not provenance.get("evidenceSummary")
            or (
                not catalog_was_reassigned
                and not expected_previous.issubset(set(provenance.get("previousValues", [])))
            )
        ):
            return []
    return credits

# Existing duplicate aliases are resolved explicitly. This map does not merge index records.
PREFERRED_COMPANY_SLUGS = {
    "505 Games": "505-games",
    "Abylight Studios": "abylight",
    "Aksys Games Localization Inc.": "aksys-games-localization",
    "ASTRAGON ENTERTAINMENT GMBH": "astragon",
    "Astragon Entertainment": "astragon",
    "astragon Entertainment": "astragon",
    "Bigben Interactive": "bigben-interactive",
    "Cavalier Game Studios Ltd": "cavalier-game-studios",
    "Curve Games": "curve-digital",
    "CyberConnect2": "cyberconnect2",
    "Delphine Software": "delphine-software-international",
    "Dolores Entertainment S.L.": "dolores-entertainment",
    "DotEmu": "dotemu",
    "Dotemu": "dotemu",
    "Eastasiasoft": "eastasiasoft",
    "eastasiasoft": "eastasiasoft",
    "EuroVideo Medien": "eurovideo",
    "FromSoftware": "fromsoftware",
    "Funbox Media": "funbox-media",
    "GAMEMILL ENTERTAINMENT": "gamemill-entertainment",
    "GameMill Entertainment": "gamemill-entertainment",
    "Game Arts": "game-arts",
    "Games Farm": "games-farm",
    "HandyGames": "handygames",
    "Harmonix Music Systems": "harmonix",
    "Idea Factory International": "idea-factory-international",
    "Infinigon Games": "infinigon",
    "Junction Point Studios": "junction-point",
    "Little Orbit": "little-orbit",
    "Lion Castle Entertainment": "lion-castle",
    "Locomalito": "locamalito",
    "MachineGames": "machinegames",
    "Mad Catz": "mad-catz",
    "Markt+Technik": "markt-technik",
    "Marvelous Europe": "marvelous-europe",
    "Maximum Entertainment": "maximum-games",
    "Microsoft Corporation": "microsoft",
    "Microsoft Studios": "microsoft-game-studios",
    "Milestone S.r.l.": "milestone-s-r-l",
    "NACON SA": "nacon",
    "NIGHTHAWK INTERACTIVE LLC": "nighthawk-interactive",
    "Nightdive Studios": "nightdive-studios",
    "OUTRIGHT GAMES LLC": "outright-games",
    "Phenomedia Publishing": "phenomedia",
    "Pix'n Love": "pix-39-n-love",
    "Pix’n Love Games": "pixn-love-games",
    "PlatinumGames": "platinumgames",
    "QUANTIC DREAM SA": "quantic-dream",
    "QUByte Interactive": "qubyte",
    "Raylight Games": "raylight-studios",
    "Square Enix": "square-enix",
    "SYSTEM 3 SOFTWARE": "system-3-software-ltd",
    "Team17": "team17",
    "THREE FIELDS ENTERTAINMENT LIMITED": "three-fields-entertainment",
    "Tivola Games": "tivola",
    "Toys For Bob": "toys-for-bob",
    "Toys for Bob": "toys-for-bob",
    "Ubisoft": "ubisoft",
    "Wildbit Studios": "wildbits-studio",
    "Z-SOFTWARE GMBH": "z-software",
    "Zoink Games": "zoink",
    "bitComposer Interactive": "bitcomposer",
    "2Awesome Studio": "2awesome-studios",
    "Chibig Studio": "chibig",
    "The Digital Lounge": "digital-lounge",
    "Platinedispositif": "platine-dispositif",
    "TINYBUILD LLC": "tinybuild",
    "tinyBuild": "tinybuild",
    "United Independent Entertainment GmbH": "united-independent-entertainment",
}
NEW_DISPLAY_NAMES = {
    "chibig": "Chibig",
    "digital-lounge": "The Digital Lounge",
    "platine-dispositif": "Platine Dispositif",
    "tinybuild": "tinyBuild",
    "united-independent-entertainment": "United Independent Entertainment",
}
WORK_KEY_OVERRIDES = {
    "ps4-gravity-rush": "gravity rush",
    "ps4-lego-the-incredibles": "lego the incredibles",
    "ps4-minecraft": "minecraft",
    "ps4-wipeout-omega-collection": "wipeout omega collection",
    "ps4-wipeout-omega-collection-only-on-playstation": "wipeout omega collection",
}
SIBLING_EVIDENCE = {
    "ps4-baldo-the-guardian-owls-collector%27s-edition": {
        "catalogId": "ps4-baldo-the-guardian-owls",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/baldo-the-guardian-owls"],
    },
    "ps4-bloodborne-goty": {
        "catalogId": "ps4-bloodborne",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/bloodborne"],
    },
    "ps4-bloodborne-ps-hits": {
        "catalogId": "ps4-bloodborne",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/bloodborne"],
    },
    "ps4-cannon-dancer-osman-collector%27s-edition": {
        "catalogId": "ps4-cannon-dancer-osman",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/cannon-dancer-osman"],
    },
    "ps4-evoland-10th-anniversary-edition": {
        "catalogId": "ps4-evoland-legendary-edition",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/evoland-legendary-edition"],
    },
    "ps4-shantae-half-genie-hero-ultimate-edition-collector%27s-edition": {
        "catalogId": "ps4-shantae-half-genie-hero-ultimate-edition",
        "urls": [
            "https://www.pricecharting.com/game/pal-playstation-4/shantae-half-genie-hero-ultimate-edition"
        ],
    },
    "ps4-zapling-bygone-deluxe-edition": {
        "catalogId": "ps4-zapling-bygone",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/zapling-bygone"],
    },
    "ps4-crash-team-rumble-deluxe-edition": {
        "catalogId": "ps4-crash-team-rumble",
        "urls": ["https://www.pricecharting.com/game/pal-playstation-4/crash-team-rumble"],
    },
}


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(path)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})
    temporary.replace(path)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_json(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def git_head() -> str:
    return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip()


def decode_display(value: Any) -> str:
    current = str(value or "")
    for _ in range(5):
        decoded = html.unescape(current)
        if decoded == current:
            break
        current = decoded
    return current.strip()


def normalized_name(value: Any) -> str:
    text = unicodedata.normalize("NFKD", decode_display(value).casefold())
    return re.sub(r"[^a-z0-9]+", "", "".join(char for char in text if not unicodedata.combining(char)))


def normalized_display(value: Any) -> str:
    return re.sub(r"\s+", " ", decode_display(value)).strip()


def work_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", decode_display(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", decode_display(value).casefold())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", text)).strip("-")


def split_companies(value: str) -> list[str]:
    return [part.strip() for part in str(value or "").split(" | ") if part.strip()]


def source_urls(row: dict[str, str]) -> list[str]:
    urls = [
        str(row.get(field) or "").strip()
        for field in ("source_1", "source_2")
        if str(row.get(field) or "").strip()
    ]
    sibling = SIBLING_EVIDENCE.get(str(row.get("catalog_id") or ""))
    return urls or list((sibling or {}).get("urls", []))


def evidence_catalog_ids(row: dict[str, str]) -> list[str]:
    sibling = SIBLING_EVIDENCE.get(str(row.get("catalog_id") or ""))
    return [sibling["catalogId"]] if sibling else []


def extract_workbook(workbook: Path) -> None:
    if sha256(workbook) != WORKBOOK_SHA256:
        raise ValueError("Workbook SHA-256 does not match the reviewed source")
    from openpyxl import load_workbook

    source = load_workbook(workbook, read_only=True, data_only=True)
    summary = source["Resumen"]
    expected_summary = {
        "B4": EXPECTED["sourceReportedReviewedRoles"],
        "B6": EXPECTED["sourceReportedWorkIdentities"],
        "E4": EXPECTED["resolvedRows"],
        "E5": 0,
        "E6": EXPECTED["conflicts"],
        "E7": EXPECTED["nonGames"],
    }
    for cell, expected in expected_summary.items():
        if summary[cell].value != expected:
            raise ValueError(f"Workbook summary {cell} changed: {summary[cell].value!r} != {expected!r}")
    outputs = {
        "Resueltos": SOURCE_RESOLVED,
        "Conflictos": SOURCE_CONFLICTS,
        "No juegos": SOURCE_NON_GAMES,
    }
    for sheet_name, output in outputs.items():
        rows = source[sheet_name].iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        extracted = [
            {header: "" if value is None else value for header, value in zip(headers, values)}
            for values in rows
            if any(value is not None for value in values)
        ]
        write_csv(output, extracted, headers)
    print(f"OK extracted reviewed workbook {workbook.name}")


def validate_sources(
    resolved: list[dict[str, str]],
    conflicts: list[dict[str, str]],
    non_games: list[dict[str, str]],
) -> None:
    if len(resolved) != EXPECTED["resolvedRows"]:
        raise ValueError(f"Expected {EXPECTED['resolvedRows']} resolved rows, found {len(resolved)}")
    if len(resolved) + len(conflicts) != EXPECTED["extractedRoleRows"]:
        raise ValueError("Extracted resolved/conflict role total changed")
    if len({row["catalog_id"] for row in resolved}) != EXPECTED["resolvedCatalogIds"]:
        raise ValueError("Resolved catalog-id count changed")
    keys = [(row["catalog_id"], row["role"]) for row in resolved]
    if len(keys) != len(set(keys)):
        raise ValueError("Resolved source has incompatible duplicate catalog_id + role keys")
    if Counter(row["role"] for row in resolved) != EXPECTED_ROLES:
        raise ValueError("Resolved role totals changed")
    if Counter(row["confidence"] for row in resolved) != {"HIGH": 1647, "VERY_HIGH": 8}:
        raise ValueError("Resolved confidence totals changed")
    if any(row["confidence"] not in {"HIGH", "VERY_HIGH"} for row in resolved):
        raise ValueError("A non-publicable confidence reached the resolved source")
    if any(not row.get("recommended_action") for row in resolved):
        raise ValueError("A resolved row has no reviewed action")
    if any(not source_urls(row) for row in resolved):
        raise ValueError("A resolved row has no direct or reviewed sibling evidence URL")
    if len(conflicts) != EXPECTED["conflicts"]:
        raise ValueError("Conflict count changed")
    if len({row["catalog_id"] for row in conflicts}) != EXPECTED["conflictCatalogIds"]:
        raise ValueError("Conflict catalog-id count changed")
    if {(row["catalog_id"], row["role"]) for row in resolved} & {
        (row["catalog_id"], row["role"]) for row in conflicts
    }:
        raise ValueError("A role is simultaneously resolved and blocked")
    if len(non_games) != EXPECTED["nonGames"]:
        raise ValueError("Non-game product count changed")
    union_ids = {
        row["catalog_id"] for row in [*resolved, *conflicts, *non_games] if row.get("catalog_id")
    }
    if len(union_ids) != EXPECTED["scopeCatalogIds"]:
        raise ValueError("Workbook catalog scope changed")


def company_lookup(
    companies: dict[str, dict[str, Any]], entities_file: dict[str, Any]
) -> dict[str, set[str]]:
    lookup: dict[str, set[str]] = defaultdict(set)
    for records in (entities_file.get("entities", {}), companies):
        for slug, entry in records.items():
            values = [slug, entry.get("name"), *(entry.get("aliasNames") or []), *(entry.get("aliasSlugs") or [])]
            for value in values:
                if value:
                    lookup[normalized_name(value)].add(slug)
    return lookup


def resolve_company_names(
    resolved: list[dict[str, str]], companies: dict[str, dict[str, Any]]
) -> dict[str, dict[str, Any]]:
    lookup = company_lookup(companies, read_json(COMPANY_ENTITIES_FILE))
    results: dict[str, dict[str, Any]] = {}
    labels = sorted({name for row in resolved for name in split_companies(row["recommended_company"])})
    generated_slugs: dict[str, str] = {}
    for label in labels:
        preferred = PREFERRED_COMPANY_SLUGS.get(label)
        matches = lookup.get(normalized_name(label), set())
        if preferred:
            slug = preferred
            method = "explicit-existing" if slug in companies else "explicit-new"
        elif len(matches) == 1:
            slug = next(iter(matches))
            method = "exact-name-or-alias"
        elif len(matches) > 1:
            raise ValueError(f"Ambiguous company label without explicit resolution: {label} -> {sorted(matches)}")
        else:
            slug = slugify(label)
            method = "new-exact-credit-entity"

        if slug in companies:
            display_name = companies[slug]["name"]
        else:
            display_name = NEW_DISPLAY_NAMES.get(slug, label)
            previous_label = generated_slugs.get(slug)
            if previous_label and previous_label != display_name:
                raise ValueError(f"Generated slug collision: {previous_label} / {display_name} -> {slug}")
            generated_slugs[slug] = display_name
        results[label] = {
            "sourceName": label,
            "slug": slug,
            "displayName": display_name,
            "resolution": method,
            "existing": slug in companies,
        }
    return results


def legacy_role_value(details: dict[str, Any] | None, role: str) -> str:
    field = "developer" if role == "DEVELOPER" else "publisher"
    value = (details or {}).get(field)
    if isinstance(value, dict):
        return decode_display(value.get("name"))
    return decode_display(value)


def build_work_identities(
    resolved: list[dict[str, str]],
    conflicts: list[dict[str, str]],
    non_games: list[dict[str, str]],
) -> dict[str, str]:
    scoped_titles: dict[str, str] = {}
    for row in [*resolved, *conflicts, *non_games]:
        catalog_id = row.get("catalog_id")
        if catalog_id:
            scoped_titles[catalog_id] = row.get("official_title") or row.get("catalog_title") or ""

    candidates: dict[str, set[str]] = defaultdict(set)
    for path in RESEARCH.glob("*.csv"):
        if path in {SOURCE_RESOLVED, SOURCE_CONFLICTS, SOURCE_NON_GAMES, DECISIONS_FILE}:
            continue
        try:
            rows = read_csv(path)
        except (UnicodeDecodeError, csv.Error):
            continue
        for row in rows:
            key = str(row.get("provisional_work_key") or "").strip()
            if not key:
                continue
            ids = str(row.get("catalog_ids") or row.get("catalog_id") or "").split(" | ")
            for catalog_id in ids:
                if catalog_id.strip() in scoped_titles:
                    candidates[catalog_id.strip()].add(key)

    output: dict[str, str] = {}
    for catalog_id, title in scoped_titles.items():
        keys = candidates.get(catalog_id, set())
        if len(keys) > 1:
            raise ValueError(f"Ambiguous previous work identity for {catalog_id}: {sorted(keys)}")
        output[catalog_id] = WORK_KEY_OVERRIDES.get(catalog_id) or (
            next(iter(keys)) if keys else work_key(title)
        )

    if len(output) != EXPECTED["scopeCatalogIds"]:
        raise ValueError("Work-identity mapping does not cover the full workbook scope")
    if len(set(output.values())) != EXPECTED["workIdentities"]:
        raise ValueError(
            f"Expected {EXPECTED['workIdentities']} work identities, found {len(set(output.values()))}"
        )
    return dict(sorted(output.items()))


def catalog_identity_matches(game: dict[str, Any], row: dict[str, str]) -> bool:
    return bool(
        game.get("platformSlug") == "ps4"
        and str(game.get("region") or "").startswith("PAL")
        and game.get("region") == row.get("catalog_region")
        and game.get("edition") == row.get("catalog_edition")
        and normalized_display(game.get("title")) == normalized_display(row.get("catalog_title"))
    )


def build_dry_run(write_artifacts: bool) -> dict[str, Any]:
    resolved = read_csv(SOURCE_RESOLVED)
    conflicts = read_csv(SOURCE_CONFLICTS)
    non_games = read_csv(SOURCE_NON_GAMES)
    validate_sources(resolved, conflicts, non_games)
    catalog = read_json(CATALOG_FILE)
    details = read_json(DETAILS_FILE)
    companies = read_json(COMPANIES_FILE)
    catalog_by_id = {row["id"]: row for row in catalog}
    if len(catalog_by_id) != len(catalog):
        raise ValueError("Catalog IDs are not unique before the dry-run")

    resolutions = resolve_company_names(resolved, companies)
    decisions: list[dict[str, Any]] = []
    for row in resolved:
        game = catalog_by_id.get(row["catalog_id"])
        current = legacy_role_value(details.get(row["catalog_id"]), row["role"])
        expected_current = decode_display(row.get("current_value"))
        recommended = decode_display(row["recommended_company"])
        stale_reasons: list[str] = []
        if not game:
            stale_reasons.append("CATALOG_ID_MISSING")
        elif not catalog_identity_matches(game, row):
            stale_reasons.append("CATALOG_IDENTITY_CHANGED")
        if current.casefold() != expected_current.casefold():
            stale_reasons.append("CURRENT_CREDIT_CHANGED")

        if stale_reasons:
            classification = "SKIP_CHANGED"
        elif not current:
            classification = "ADD"
        elif current.casefold() == recommended.casefold():
            classification = "CONFIRM"
        else:
            classification = "REPLACE"

        names = split_companies(row["recommended_company"])
        decision = dict(row)
        decision.update(
            {
                "classification": classification,
                "previous_value": current,
                "company_slugs": " | ".join(resolutions[name]["slug"] for name in names),
                "company_display_names": " | ".join(
                    resolutions[name]["displayName"] for name in names
                ),
                "resolved_evidence_urls": " | ".join(source_urls(row)),
                "evidence_catalog_ids": " | ".join(evidence_catalog_ids(row)),
                "stale_reason": " | ".join(stale_reasons),
                "reviewed_at": REVIEWED_AT,
                "review_batch": BATCH_ID,
            }
        )
        decisions.append(decision)

    action_counts = Counter(row["classification"] for row in decisions)
    for action in EXPECTED_ACTIONS:
        action_counts.setdefault(action, 0)
    if dict(action_counts) != EXPECTED_ACTIONS:
        raise ValueError(f"Dry-run action totals changed: {dict(action_counts)}")

    non_game_decisions = []
    for row in non_games:
        game = catalog_by_id.get(row["catalog_id"])
        if not game:
            status = "SKIP_CHANGED"
        elif game.get("platformSlug") != "ps4" or not str(game.get("region") or "").startswith("PAL"):
            status = "SKIP_CHANGED"
        elif game.get("listingStatus") == "excluded":
            status = "CONFIRM_EXCLUDED"
        else:
            status = "EXCLUDE"
        non_game_decisions.append({**row, "classification": status, "review_batch": BATCH_ID})
    if Counter(row["classification"] for row in non_game_decisions) != {
        "CONFIRM_EXCLUDED": 2,
        "EXCLUDE": 4,
    }:
        raise ValueError("Non-game dry-run classifications changed")

    work_identities = build_work_identities(resolved, conflicts, non_games)
    report = {
        "schemaVersion": 1,
        "batchId": BATCH_ID,
        "reviewedAt": REVIEWED_AT,
        "sourceWorkbook": {"name": WORKBOOK_NAME, "sha256": WORKBOOK_SHA256},
        "git": {"stackedBaseHead": git_head(), "headAtGeneration": git_head()},
        "writesPerformed": False,
        "sourceArtifacts": {
            str(path.relative_to(ROOT)): sha256(path)
            for path in (SOURCE_RESOLVED, SOURCE_CONFLICTS, SOURCE_NON_GAMES)
        },
        "summary": {
            "sourceReportedReviewedRoles": EXPECTED["sourceReportedReviewedRoles"],
            "extractedRoleRows": EXPECTED["extractedRoleRows"],
            "sourceRoleCountDelta": EXPECTED["sourceReportedReviewedRoles"]
            - EXPECTED["extractedRoleRows"],
            "resolvedRows": len(decisions),
            "resolvedCatalogIds": len({row["catalog_id"] for row in decisions}),
            "actions": dict(action_counts),
            "roles": dict(Counter(row["role"] for row in decisions)),
            "multiCompanyCredits": sum(" | " in row["recommended_company"] for row in decisions),
            "conflictsBlocked": len(conflicts),
            "nonGameProducts": len(non_game_decisions),
            "newlyExcluded": sum(row["classification"] == "EXCLUDE" for row in non_game_decisions),
            "alreadyExcluded": sum(
                row["classification"] == "CONFIRM_EXCLUDED" for row in non_game_decisions
            ),
            "companyLabels": len(resolutions),
            "newCompanyEntities": len(
                {value["slug"] for value in resolutions.values() if not value["existing"]}
            ),
            "workMappedCatalogIds": len(work_identities),
            "sourceReportedWorkIdentities": EXPECTED["sourceReportedWorkIdentities"],
            "workIdentities": len(set(work_identities.values())),
            "workIdentityAdjustments": [
                {
                    "workKey": "wipeout omega collection",
                    "catalogIds": [
                        "ps4-wipeout-omega-collection",
                        "ps4-wipeout-omega-collection-only-on-playstation",
                    ],
                    "reason": "ONLY_ON_PLAYSTATION_IS_A_MARKETING_VARIANT_OF_THE_SAME_WORK",
                }
            ],
        },
        "catalogBefore": {
            "rows": len(catalog),
            "uniqueIds": len(catalog_by_id),
            "sha256": sha256(CATALOG_FILE),
        },
        "detailsBeforeSha256": sha256(DETAILS_FILE),
        "companiesBefore": {"rows": len(companies), "sha256": sha256(COMPANIES_FILE)},
        "scopeChecks": {
            "allResolvedIdsExist": all(row["catalog_id"] in catalog_by_id for row in decisions),
            "allResolvedRowsArePs4Pal": all(
                catalog_by_id[row["catalog_id"]]["platformSlug"] == "ps4"
                and str(catalog_by_id[row["catalog_id"]]["region"]).startswith("PAL")
                for row in decisions
            ),
            "uniqueResolvedCatalogIdRole": len(decisions)
            == len({(row["catalog_id"], row["role"]) for row in decisions}),
            "noResolvedConflictRoleOverlap": not {
                (row["catalog_id"], row["role"]) for row in decisions
            }
            & {(row["catalog_id"], row["role"]) for row in conflicts},
            "noSkipChanged": action_counts["SKIP_CHANGED"] == 0,
            "allConflictsBlocked": len(conflicts) == EXPECTED["conflicts"],
        },
        "verification": {
            "semanticComparator": "PENDING",
            "typecheck": "PENDING",
            "lint": "PENDING",
            "unitTests": "PENDING",
            "collectorControls": "PENDING",
            "affiliateOffersV1": "PENDING",
            "build": "PENDING",
            "previewQa": "PENDING",
        },
    }
    if not all(report["scopeChecks"].values()):
        raise ValueError(f"Dry-run scope guard failed: {report['scopeChecks']}")

    if write_artifacts:
        write_csv(DECISIONS_FILE, decisions, list(decisions[0]))
        write_json(
            BLOCKED_FILE,
            {
                "schemaVersion": 1,
                "batchId": BATCH_ID,
                "status": "BLOCKED",
                "count": len(conflicts),
                "rows": conflicts,
            },
        )
        write_csv(NON_GAME_DECISIONS_FILE, non_game_decisions, list(non_game_decisions[0]))
        write_csv(
            COMPANY_RESOLUTION_FILE,
            [resolutions[label] for label in sorted(resolutions)],
            ["sourceName", "slug", "displayName", "resolution", "existing"],
        )
        write_json(REPORT_FILE, report)
        REPORT_MD_FILE.write_text(render_markdown(report), encoding="utf-8")
    return {"report": report, "decisions": decisions, "workIdentities": work_identities}


def new_detail() -> dict[str, Any]:
    return {
        "year": None,
        "releaseDate": None,
        "reference": None,
        "players": None,
        "support": None,
        "developer": None,
        "publisher": None,
        "genres": [],
        "series": None,
        "museumPath": None,
        "pcProductId": None,
        "ean": None,
        "sources": {},
        "fieldSources": {},
        "fetchedAt": REVIEWED_AT_TIMESTAMP,
        "mergedAt": REVIEWED_AT_TIMESTAMP,
    }


def provenance(row: dict[str, str]) -> dict[str, Any]:
    previous = split_companies(row.get("previous_value", ""))
    value = {
        "source": "research",
        "evidenceUrls": source_urls(row),
        "evidenceSummary": row["evidence_summary"],
        "reviewedAt": REVIEWED_AT,
        "reviewBatch": BATCH_ID,
    }
    if previous:
        value["previousValues"] = previous
    evidence_ids = evidence_catalog_ids(row)
    if evidence_ids:
        value["evidenceCatalogIds"] = evidence_ids
    return value


def detail_entity(slug: str, companies: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return {
        "name": companies[slug]["name"],
        "slug": slug,
        "museumPath": None,
        "pcPath": None,
        "source": "research",
    }


def ensure_company(
    companies: dict[str, dict[str, Any]], slug: str, display_name: str
) -> dict[str, Any]:
    if slug not in companies:
        companies[slug] = {
            "name": display_name,
            "slug": slug,
            "museumPath": "",
            "gameIds": [],
            "byPlatform": {},
            "gameCount": 0,
            "asDeveloper": [],
            "asPublisher": [],
            "asDigitalPublisher": [],
            "asPhysicalPublisherOrDistributor": [],
        }
    return companies[slug]


def remove_role_credit(
    companies: dict[str, dict[str, Any]], catalog_id: str, role_field: str, slug: str
) -> set[str]:
    changed: set[str] = set()
    entry = companies.get(slug)
    if not entry:
        return changed
    if catalog_id in entry.get(role_field, []):
        entry[role_field] = [item for item in entry[role_field] if item != catalog_id]
        remaining_role_fields = (
            "asDeveloper",
            "asPublisher",
            "asDigitalPublisher",
            "asPhysicalPublisherOrDistributor",
        )
        if not any(catalog_id in entry.get(field, []) for field in remaining_role_fields):
            entry["gameIds"] = [item for item in entry.get("gameIds", []) if item != catalog_id]
        changed.add(slug)
    return changed


def add_role_credit(
    companies: dict[str, dict[str, Any]], catalog_id: str, role_field: str, slug: str
) -> None:
    entry = companies[slug]
    if catalog_id not in entry.setdefault(role_field, []):
        entry[role_field].append(catalog_id)
    if catalog_id not in entry.setdefault("gameIds", []):
        entry["gameIds"].append(catalog_id)


def refresh_company_entry(entry: dict[str, Any], catalog_by_id: dict[str, dict[str, Any]]) -> None:
    role_fields = (
        "asDeveloper",
        "asPublisher",
        "asDigitalPublisher",
        "asPhysicalPublisherOrDistributor",
    )
    for field in role_fields:
        if field in entry:
            entry[field] = list(dict.fromkeys(entry[field]))
    entry["gameIds"] = [
        catalog_id
        for catalog_id in dict.fromkeys(entry.get("gameIds", []))
        if catalog_id in catalog_by_id
    ]
    by_platform = Counter(catalog_by_id[catalog_id]["platformSlug"] for catalog_id in entry["gameIds"])
    entry["byPlatform"] = dict(sorted(by_platform.items()))
    entry["gameCount"] = len(entry["gameIds"])


def protected_catalog_entry(game: dict[str, Any]) -> dict[str, Any]:
    allowed = {"listingStatus", "catalogKind", "hardwareMetadata", "excludeCategory", "excludeReason"}
    return {key: value for key, value in game.items() if key not in allowed}


def protected_catalog_hash(catalog: list[dict[str, Any]]) -> str:
    return sha256_json([protected_catalog_entry(game) for game in catalog])


def counts_in_existing_order(current: dict[str, Any], counts: Counter[str]) -> dict[str, int]:
    ordered = {key: counts[key] for key in current if key in counts}
    for key in sorted(set(counts) - set(ordered)):
        ordered[key] = counts[key]
    return ordered


def recompute_catalog_meta(
    catalog: list[dict[str, Any]], details: dict[str, dict[str, Any]], companies: dict[str, Any]
) -> tuple[dict[str, Any], dict[str, Any]]:
    meta = read_json(META_FILE)
    curation = read_json(CURATION_FILE)
    def is_listed_game(game: dict[str, Any]) -> bool:
        return (
            game.get("listingStatus") != "excluded"
            and game.get("catalogKind", "game") == "game"
        )

    listed = [game for game in catalog if is_listed_game(game)]
    excluded = [game for game in catalog if not is_listed_game(game)]
    listed_by_platform = Counter(game["platformSlug"] for game in listed)
    excluded_by_platform = Counter(game["platformSlug"] for game in excluded)
    by_category = Counter(game.get("excludeCategory") or "other" for game in excluded)

    meta.update(
        {
            "catalogListed": len(listed),
            "listedByPlatform": counts_in_existing_order(
                meta.get("listedByPlatform", {}), listed_by_platform
            ),
            "catalogExcluded": len(excluded),
            "catalogTotal": len(catalog),
            "excludedByPlatform": counts_in_existing_order(
                meta.get("excludedByPlatform", {}), excluded_by_platform
            ),
            "curationByCategory": counts_in_existing_order(
                meta.get("curationByCategory", {}), by_category
            ),
            "coversListed": sum(bool(game.get("coverUrl")) for game in listed),
            "coversListedPct": round(
                100 * sum(bool(game.get("coverUrl")) for game in listed) / max(len(listed), 1), 1
            ),
            "gamesWithDetails": len(details),
            "indexCompanies": len(companies),
        }
    )
    curation.update(
        {
            "total": len(catalog),
            "listed": len(listed),
            "excluded": len(excluded),
            "byCategory": counts_in_existing_order(curation.get("byCategory", {}), by_category),
            "listedByPlatform": counts_in_existing_order(
                curation.get("listedByPlatform", {}), listed_by_platform
            ),
            "excludedByPlatform": counts_in_existing_order(
                curation.get("excludedByPlatform", {}), excluded_by_platform
            ),
        }
    )
    return meta, curation


def update_manifest(
    manifest: dict[str, Any], before_hashes: dict[str, str], after_hashes: dict[str, str]
) -> dict[str, Any]:
    updated = copy.deepcopy(manifest)
    files = {
        path: {"before": before_hashes[path], "after": after_hashes[path]}
        for path in before_hashes
    }
    updates = [
        item
        for item in updated.setdefault("protectedFileHashUpdates", [])
        if item.get("batchId") != BATCH_ID
    ]
    updates.append({"batchId": BATCH_ID, "reviewedAt": REVIEWED_AT, "files": files})
    updated["protectedFileHashUpdates"] = updates
    for path, digest in after_hashes.items():
        updated.setdefault("protectedFileHashes", {})[path] = digest
    return updated


def apply_non_games(
    catalog_by_id: dict[str, dict[str, Any]], companies: dict[str, dict[str, Any]]
) -> tuple[list[str], set[str]]:
    rows = read_csv(SOURCE_NON_GAMES)
    changed_ids: list[str] = []
    affected_companies: set[str] = set()
    kind_map = {
        "ACCESSORY": "accessory",
        "CONSOLE": "console",
        "CONSOLE_BUNDLE": "console_bundle",
        "SUBSCRIPTION": "subscription",
    }
    category_map = {
        "ACCESSORY": "accessory",
        "CONSOLE": "hardware",
        "CONSOLE_BUNDLE": "hardware",
        "SUBSCRIPTION": "other",
    }
    for row in rows:
        game = catalog_by_id[row["catalog_id"]]
        game.update(
            {
                "listingStatus": "excluded",
                "catalogKind": kind_map[row["product_type"]],
                "excludeCategory": category_map[row["product_type"]],
                "excludeReason": BATCH_ID,
                "hardwareMetadata": {
                    "brand": row.get("brand") or None,
                    "manufacturer": row.get("manufacturer") or None,
                    "model": row.get("model") or None,
                    "ean": str(row.get("ean") or "").lstrip("'") or None,
                    "sourceUrls": source_urls(row),
                    "reviewedAt": REVIEWED_AT,
                    "reviewBatch": BATCH_ID,
                },
            }
        )
        changed_ids.append(row["catalog_id"])
        for slug, entry in companies.items():
            touched = False
            for field in (
                "gameIds",
                "asDeveloper",
                "asPublisher",
                "asDigitalPublisher",
                "asPhysicalPublisherOrDistributor",
            ):
                if row["catalog_id"] in entry.get(field, []):
                    entry[field] = [item for item in entry[field] if item != row["catalog_id"]]
                    touched = True
            if touched:
                affected_companies.add(slug)
    return changed_ids, affected_companies


def run_apply() -> dict[str, Any]:
    dry = build_dry_run(write_artifacts=True)
    report = dry["report"]
    decisions = dry["decisions"]
    if report["summary"]["actions"]["SKIP_CHANGED"]:
        raise ValueError("Dry-run contains SKIP_CHANGED rows; no data was written")

    catalog = read_json(CATALOG_FILE)
    details = read_json(DETAILS_FILE)
    companies = read_json(COMPANIES_FILE)
    catalog_before = copy.deepcopy(catalog)
    details_before = copy.deepcopy(details)
    companies_before = copy.deepcopy(companies)
    catalog_by_id = {game["id"]: game for game in catalog}
    resolutions = {row["sourceName"]: row for row in read_csv(COMPANY_RESOLUTION_FILE)}
    affected_company_slugs: set[str] = set()

    for resolution in resolutions.values():
        if resolution["slug"] not in companies:
            ensure_company(companies, resolution["slug"], resolution["displayName"])
            affected_company_slugs.add(resolution["slug"])

    for row in decisions:
        catalog_id = row["catalog_id"]
        role = row["role"]
        role_name = ROLE_NAMES[role]
        index_field = ROLE_INDEX_FIELDS[role]
        slugs = row["company_slugs"].split(" | ")
        detail = details.setdefault(catalog_id, new_detail())
        old_field = "developer" if role == "DEVELOPER" else "publisher"
        old_entity = detail.get(old_field)

        explicit = [credit for credit in detail.get("companyCredits", []) if credit.get("role") != role_name]
        for slug in slugs:
            explicit.append(
                {
                    "role": role_name,
                    "company": detail_entity(slug, companies),
                    "provenance": provenance(row),
                }
            )
            add_role_credit(companies, catalog_id, index_field, slug)
            affected_company_slugs.add(slug)
        detail["companyCredits"] = explicit

        if role in {"DEVELOPER", "PUBLISHER"} and row["classification"] in {"ADD", "REPLACE"}:
            if (
                isinstance(old_entity, dict)
                and old_entity.get("slug")
                and old_entity["slug"] not in slugs
            ):
                affected_company_slugs |= remove_role_credit(
                    companies, catalog_id, index_field, old_entity["slug"]
                )
            detail[old_field] = detail_entity(slugs[0], companies)
            detail.setdefault("fieldSources", {})[old_field] = "research"
            detail.setdefault("fieldProvenance", {})[old_field] = provenance(row)

    non_game_ids, non_game_companies = apply_non_games(catalog_by_id, companies)
    affected_company_slugs |= non_game_companies
    for slug in affected_company_slugs:
        refresh_company_entry(companies[slug], catalog_by_id)

    work_index = {
        "schemaVersion": 1,
        "sourceBatch": BATCH_ID,
        "catalogIdToWorkKey": dry["workIdentities"],
    }
    meta, curation = recompute_catalog_meta(catalog, details, companies)

    allowed_catalog_ids = set(non_game_ids)
    changed_catalog_ids = {
        after["id"]
        for before, after in zip(catalog_before, catalog)
        if before != after
    }
    if changed_catalog_ids != allowed_catalog_ids:
        raise ValueError(f"Unexpected catalog changes: {sorted(changed_catalog_ids ^ allowed_catalog_ids)}")
    if protected_catalog_hash(catalog_before) != protected_catalog_hash(catalog):
        raise ValueError("A protected catalog field changed")

    target_detail_ids = {row["catalog_id"] for row in decisions}
    changed_detail_ids = {
        catalog_id
        for catalog_id in details_before.keys() | details.keys()
        if details_before.get(catalog_id) != details.get(catalog_id)
    }
    if not changed_detail_ids <= target_detail_ids:
        raise ValueError(f"Details propagated outside scope: {sorted(changed_detail_ids - target_detail_ids)}")
    changed_company_slugs = {
        slug
        for slug in companies_before.keys() | companies.keys()
        if companies_before.get(slug) != companies.get(slug)
    }
    if not changed_company_slugs <= affected_company_slugs:
        raise ValueError(
            f"Company index changed outside scope: {sorted(changed_company_slugs - affected_company_slugs)}"
        )

    before_hashes = {
        "data/game-details.json": sha256(DETAILS_FILE),
        "data/index/companies.json": sha256(COMPANIES_FILE),
    }
    after_hashes = {
        "data/game-details.json": sha256_json(details),
        "data/index/companies.json": sha256_json(companies),
    }
    manifests = {
        path: update_manifest(read_json(path), before_hashes, after_hashes) for path in MANIFEST_FILES
    }

    report.update(
        {
            "writesPerformed": True,
            "catalogAfter": {
                "rows": len(catalog),
                "uniqueIds": len({game["id"] for game in catalog}),
                "sha256": sha256_json(catalog),
                "protectedFieldsSha256Before": protected_catalog_hash(catalog_before),
                "protectedFieldsSha256After": protected_catalog_hash(catalog),
                "changedIds": sorted(changed_catalog_ids),
                "listed": meta["catalogListed"],
                "excluded": meta["catalogExcluded"],
            },
            "detailsAfter": {
                "sha256": after_hashes["data/game-details.json"],
                "changedCatalogIds": len(changed_detail_ids),
                "changedCatalogIdList": sorted(changed_detail_ids),
                "createdRecords": len(set(details) - set(details_before)),
            },
            "companiesAfter": {
                "rows": len(companies),
                "sha256": after_hashes["data/index/companies.json"],
                "created": len(set(companies) - set(companies_before)),
                "createdSlugs": sorted(set(companies) - set(companies_before)),
                "affected": len(affected_company_slugs),
                "changed": len(changed_company_slugs),
                "changedSlugs": sorted(changed_company_slugs),
            },
            "scopeChecksAfter": {
                "catalogRowsAndIdsPreserved": len(catalog) == len(catalog_before)
                == len({game["id"] for game in catalog}),
                "onlySixCatalogRecordsChanged": changed_catalog_ids == allowed_catalog_ids,
                "idsUrlsSlugsRegionsPlatformsEditionsCoversPricesFranchisesUnchanged": protected_catalog_hash(
                    catalog_before
                )
                == protected_catalog_hash(catalog),
                "detailsOnlyChangedForResolvedIds": changed_detail_ids <= target_detail_ids,
                "companyIndexOnlyChangedForAffectedEntities": changed_company_slugs
                <= affected_company_slugs,
                "allConflictsRemainBlocked": len(read_json(BLOCKED_FILE)["rows"])
                == EXPECTED["conflicts"],
                "noRegionalPlatformOrEditionPropagation": all(
                    catalog_by_id[catalog_id]["platformSlug"] == "ps4"
                    and str(catalog_by_id[catalog_id]["region"]).startswith("PAL")
                    for catalog_id in target_detail_ids
                ),
                "sixNonGamesExcluded": all(
                    catalog_by_id[catalog_id]["listingStatus"] == "excluded"
                    and catalog_by_id[catalog_id].get("catalogKind") != "game"
                    for catalog_id in non_game_ids
                ),
                "workIdentityCountPreserved": len(set(dry["workIdentities"].values()))
                == EXPECTED["workIdentities"],
            },
        }
    )
    if not all(report["scopeChecksAfter"].values()):
        raise ValueError(f"Post-apply scope guard failed: {report['scopeChecksAfter']}")

    write_json(CATALOG_FILE, catalog)
    write_json(DETAILS_FILE, details)
    write_json(COMPANIES_FILE, companies)
    write_json(WORK_IDENTITIES_FILE, work_index)
    write_json(META_FILE, meta)
    write_json(CURATION_FILE, curation)
    for path, manifest in manifests.items():
        write_json(path, manifest)
    subprocess.run([sys.executable, str(VERIFIED_INDEX_BUILDER), "--write"], cwd=ROOT, check=True)
    write_json(REPORT_FILE, report)
    REPORT_MD_FILE.write_text(render_markdown(report), encoding="utf-8")
    return report


def render_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    actions = summary["actions"]
    verification = report["verification"]
    catalog_before = report["catalogBefore"]
    catalog_after = report.get("catalogAfter")
    companies_before = report["companiesBefore"]
    companies_after = report.get("companiesAfter")
    lines = [
        "# PS4 PAL - repaso rapido de creditos por funcion",
        "",
        f"Lote: `{BATCH_ID}`",
        f"Fuente: `{WORKBOOK_NAME}` (`{WORKBOOK_SHA256}`)",
        "",
        "## Dry-run",
        "",
        f"- Roles que declara el resumen del libro: {summary['sourceReportedReviewedRoles']}.",
        f"- Filas de rol realmente extraidas: {summary['extractedRoleRows']} (diferencia de origen: {summary['sourceRoleCountDelta']}).",
        f"- ADD: {actions['ADD']}",
        f"- REPLACE: {actions['REPLACE']}",
        f"- CONFIRM: {actions['CONFIRM']}",
        f"- SKIP_CHANGED: {actions['SKIP_CHANGED']}",
        f"- Conflictos bloqueados: {summary['conflictsBlocked']}",
        f"- Productos no juego: {summary['nonGameProducts']}",
        "",
        "## Aplicacion",
        "",
        f"- Filas de rol resueltas: {summary['resolvedRows']} sobre {summary['resolvedCatalogIds']} fichas.",
        f"- Desarrolladora: {summary['roles']['DEVELOPER']}.",
        f"- Publicadora generica: {summary['roles']['PUBLISHER']}.",
        f"- Editora digital: {summary['roles']['DIGITAL_PUBLISHER']}.",
        f"- Editora o distribuidora fisica: {summary['roles']['PHYSICAL_PUBLISHER_OR_DISTRIBUTOR']}.",
        f"- Creditos adicionales por co-desarrollo: {summary['multiCompanyCredits']}.",
        f"- Entidades de compania nuevas: {summary['newCompanyEntities']}.",
        f"- Productos excluidos ahora: {summary['newlyExcluded']}; ya excluidos: {summary['alreadyExcluded']}.",
        "",
        "## Semantica",
        "",
        "Los creditos de desarrollo, publicacion general, edicion digital y edicion o distribucion fisica se almacenan por separado. Los creditos multiples conservan cada entidad.",
        f"La identidad editorial cubre {summary['workMappedCatalogIds']} fichas con {summary['workIdentities']} obras explicitas.",
        f"El libro fuente indicaba {summary['sourceReportedWorkIdentities']}; se corrige una unidad porque Wipeout Omega Collection y su variante Only On PlayStation son la misma obra.",
        "El total general del Resumen declara diez roles mas que las hojas Resueltos y Conflictos. La importacion conserva la discrepancia y no inventa filas ausentes.",
        "",
        "## Invariantes",
        "",
        f"- Catalogo: {catalog_before['rows']} -> {catalog_after['rows'] if catalog_after else 'sin escritura'} fichas.",
        f"- IDs unicos: {catalog_before['uniqueIds']} -> {catalog_after['uniqueIds'] if catalog_after else 'sin escritura'}.",
        f"- Companias: {companies_before['rows']} -> {companies_after['rows'] if companies_after else 'sin escritura'}.",
        "- IDs, URLs, slugs, regiones, plataformas, ediciones, portadas, precios y franquicias: sin cambios.",
        "- Los 95 conflictos siguen bloqueados y no existe propagacion a otras regiones o plataformas.",
        "",
        "## Verificacion",
        "",
        f"- Comparador semantico: {verification['semanticComparator']}.",
        f"- Typecheck: {verification['typecheck']}.",
        f"- Lint: {verification['lint']}.",
        f"- Pruebas unitarias: {verification['unitTests']}.",
        f"- Controles de recolectores: {verification['collectorControls']}.",
        f"- Afiliacion: {verification['affiliateOffersV1']}.",
        f"- Build: {verification['build']}.",
        f"- QA Preview: {verification['previewQa']}.",
        "",
        "## Estado",
        "",
        "Aplicacion realizada." if report.get("writesPerformed") else "Dry-run reproducible; sin mutaciones de catalogo.",
        "",
        "Quality y QA Preview se documentan tambien en la PR para su HEAD remoto exacto.",
        "",
    ]
    return "\n".join(lines)


def check_committed() -> None:
    resolved = read_csv(SOURCE_RESOLVED)
    conflicts = read_csv(SOURCE_CONFLICTS)
    non_games = read_csv(SOURCE_NON_GAMES)
    decisions = read_csv(DECISIONS_FILE)
    validate_sources(resolved, conflicts, non_games)
    action_counts = Counter(row["classification"] for row in decisions)
    for action in EXPECTED_ACTIONS:
        action_counts.setdefault(action, 0)
    if dict(action_counts) != EXPECTED_ACTIONS:
        raise ValueError("Committed action totals changed")
    report = read_json(REPORT_FILE)
    catalog = read_json(CATALOG_FILE)
    details = read_json(DETAILS_FILE)
    companies = read_json(COMPANIES_FILE)
    work_index = read_json(WORK_IDENTITIES_FILE)
    catalog_by_id = {game["id"]: game for game in catalog}
    if len(catalog_by_id) != len(catalog):
        raise ValueError("Catalog IDs are not unique")
    if report.get("writesPerformed") is not True:
        raise ValueError("Committed report is not the applied report")
    if not all(report["scopeChecksAfter"].values()):
        raise ValueError("A committed post-apply scope check failed")
    batch_identities = build_work_identities(resolved, conflicts, non_games)
    for catalog_id, work_key_value in batch_identities.items():
        effective_id = SUCCESSOR_CATALOG_IDS.get(catalog_id, catalog_id)
        if work_index["catalogIdToWorkKey"].get(effective_id) != work_key_value:
            raise ValueError(f"Committed batch work identity changed: {effective_id}")
    if read_json(BLOCKED_FILE)["count"] != EXPECTED["conflicts"]:
        raise ValueError("Blocked conflicts changed")

    for row in decisions:
        effective_catalog_id = SUCCESSOR_CATALOG_IDS.get(
            row["catalog_id"], row["catalog_id"]
        )
        role_name = ROLE_NAMES[row["role"]]
        expected_slugs = row["company_slugs"].split(" | ")
        role_credits = [
            credit
            for credit in details[effective_catalog_id].get("companyCredits", [])
            if credit.get("role") == role_name
            and credit.get("provenance", {}).get("reviewBatch") == BATCH_ID
        ]
        successor_credits = allowed_successor_role_credits(
            row, details[effective_catalog_id]
        )
        if (
            [credit["company"]["slug"] for credit in role_credits] != expected_slugs
            and not successor_credits
        ):
            raise ValueError(f"Role credit mismatch for {row['catalog_id']}:{row['role']}")
        effective_credits = role_credits or successor_credits
        if any(
            credit["provenance"].get("reviewedAt") != REVIEWED_AT
            or not credit["provenance"].get("evidenceSummary")
            or not credit["provenance"].get("evidenceUrls")
            for credit in effective_credits
        ):
            raise ValueError(f"Incomplete provenance for {row['catalog_id']}:{row['role']}")
        for slug in [credit["company"]["slug"] for credit in effective_credits]:
            if effective_catalog_id not in companies[slug][ROLE_INDEX_FIELDS[row["role"]]]:
                raise ValueError(f"Company index missing {effective_catalog_id}:{slug}")

    conflict_keys = {(row["catalog_id"], ROLE_NAMES[row["role"]]) for row in conflicts}
    for catalog_id, role_name in conflict_keys:
        if any(
            credit.get("role") == role_name
            and credit.get("provenance", {}).get("reviewBatch") == BATCH_ID
            for credit in details.get(catalog_id, {}).get("companyCredits", [])
        ):
            raise ValueError(f"Blocked conflict was published: {catalog_id}:{role_name}")

    for row in non_games:
        game = catalog_by_id[row["catalog_id"]]
        if game["listingStatus"] != "excluded" or game.get("catalogKind", "game") == "game":
            raise ValueError(f"Non-game still public: {row['catalog_id']}")
        if any(row["catalog_id"] in company.get("gameIds", []) for company in companies.values()):
            raise ValueError(f"Non-game still indexed as company game: {row['catalog_id']}")

    print(
        "OK PS4 PAL rapid review: "
        f"{len(decisions)} role rows, {EXPECTED['conflicts']} conflicts blocked, "
        f"{EXPECTED['nonGames']} non-games excluded, {len(catalog)} unique catalog IDs"
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--extract-workbook", type=Path)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    mode.add_argument("--check", action="store_true")
    args = parser.parse_args()
    if args.extract_workbook:
        extract_workbook(args.extract_workbook.resolve())
    elif args.dry_run:
        result = build_dry_run(write_artifacts=True)["report"]
        print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
    elif args.apply:
        result = run_apply()
        print(json.dumps(result["summary"], ensure_ascii=False, indent=2))
    else:
        check_committed()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
